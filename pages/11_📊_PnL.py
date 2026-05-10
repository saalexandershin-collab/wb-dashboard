import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import calendar
import io
from datetime import date as date_cls
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.db.models import init_db, get_session_factory
from src.db.repository import FinancialReportRepository, OzonTransactionRepository
from src.auth import require_role

require_role(["admin"])
st.title("📊 Сводный P&L отчёт")

if "database" not in st.secrets:
    st.error("Настройте базу данных в разделе ⚙️ Настройки.")
    st.stop()

DB_URL = st.secrets["database"]["url"]

# ── Период ────────────────────────────────────────────────────────────────────
today = date_cls.today()
col_y, col_m = st.sidebar.columns(2)
year  = col_y.selectbox("Год",   list(range(today.year, today.year - 3, -1)), index=0)
month = col_m.selectbox("Месяц", list(range(1, 13)), index=today.month - 1,
                        format_func=lambda m: calendar.month_name[m])
st.sidebar.markdown("---")

# ── Налоговые ставки ──────────────────────────────────────────────────────────
st.sidebar.markdown("### Налоги")
nds_pct  = st.sidebar.number_input("НДС (%)",               value=5.0, min_value=0.0, max_value=30.0, step=0.5)
usn_pct  = st.sidebar.number_input("УСН (%)",               value=1.0, min_value=0.0, max_value=15.0, step=0.5)
ins_pct  = st.sidebar.number_input("Страховые взносы (%)",  value=1.0, min_value=0.0, max_value=10.0, step=0.5)
total_tax_pct = nds_pct + usn_pct + ins_pct
st.sidebar.caption(f"Итого налоги: **{total_tax_pct:.1f}%**")
st.sidebar.markdown("---")
st.sidebar.caption(f"Период: {calendar.month_name[month]} {year}")


# ── Загрузка данных ───────────────────────────────────────────────────────────
@st.cache_data(ttl=300, show_spinner="Загружаю данные...")
def load_data(db_url: str, year: int, month: int):
    engine = init_db(db_url)
    Session = get_session_factory(engine)
    with Session() as session:
        wb  = FinancialReportRepository().get_by_month(session, year, month)
        oz  = OzonTransactionRepository().get_by_month(session, year, month)
    return wb, oz


df_wb, df_oz = load_data(DB_URL, year, month)

s = lambda col: float(col.fillna(0).sum())


# ── WB: чистая выплата ────────────────────────────────────────────────────────
# ppvz_for_pay = выплата за продажи/возвраты (уже за вычетом комиссии WB + эквайринга)
# delivery_rub, storage_fee, penalty хранятся как ПОЛОЖИТЕЛЬНЫЕ числа → вычитаем
wb_empty = df_wb.empty
if not wb_empty:
    sales_mask   = df_wb["doc_type_name"] == "Продажа"
    returns_mask = df_wb["doc_type_name"] == "Возврат"

    wb_ppvz_sales   = s(df_wb.loc[sales_mask,   "ppvz_for_pay"])   # + выплата за продажи
    wb_ppvz_returns = s(df_wb.loc[returns_mask,  "ppvz_for_pay"])   # + / − за возвраты
    wb_ppvz         = wb_ppvz_sales + wb_ppvz_returns

    wb_delivery  = s(df_wb["delivery_rub"])      # + в базе, но это РАСХОД (логистика)
    wb_storage   = s(df_wb["storage_fee"])        # + в базе, но это РАСХОД (хранение)
    wb_penalties = s(df_wb["penalty"])            # + в базе, но это РАСХОД (штрафы)
    wb_other     = s(df_wb["additional_payment"]) # ± прочее (хранится со знаком)

    wb_qty_sold     = int(df_wb.loc[sales_mask,   "quantity"].clip(lower=0).sum())
    wb_qty_returned = int(df_wb.loc[returns_mask,  "quantity"].fillna(0).abs().sum())

    # Фактическая выплата WB = ppvz − логистика − хранение − штрафы ± прочее
    wb_net = wb_ppvz - wb_delivery - wb_storage - wb_penalties + wb_other
else:
    wb_ppvz = wb_ppvz_sales = wb_ppvz_returns = 0.0
    wb_delivery = wb_storage = wb_penalties = wb_other = wb_net = 0.0
    wb_qty_sold = wb_qty_returned = 0


# ── Ozon: чистая выплата ──────────────────────────────────────────────────────
# amount хранится со знаком: + для продаж, − для возвратов и удержаний
RETURN_OPS_SET = {
    "OperationReturnGoodsFBSofRMS", "OperationReturnGoodsFBO",
    "OperationItemReturn", "OperationReturnGoodsFBSofMerchant",
    "OperationReturnGoodsFBSofRMSReject", "OperationLagTimeReturnAfterDelivery",
}
oz_empty = df_oz.empty
if not oz_empty:
    oz_sales_mask   = df_oz["operation_type"] == "OperationAgentDeliveredToCustomer"
    oz_returns_mask = df_oz["operation_type"].isin(RETURN_OPS_SET) | \
                      df_oz["operation_type"].str.contains("return|Return", na=False)
    oz_other_mask   = ~oz_sales_mask & ~oz_returns_mask

    oz_sales_amount   = s(df_oz.loc[oz_sales_mask,   "amount"])  # + выплата за доставленные
    oz_returns_amount = s(df_oz.loc[oz_returns_mask,  "amount"])  # − удержание за возвраты
    oz_other_amount   = s(df_oz.loc[oz_other_mask,    "amount"])  # − прочие удержания

    oz_qty_sold     = int(oz_sales_mask.sum())
    oz_qty_returned = int(oz_returns_mask.sum())

    # Нетто Ozon = сумма ВСЕХ транзакций
    oz_net = s(df_oz["amount"])
else:
    oz_sales_amount = oz_returns_amount = oz_other_amount = oz_net = 0.0
    oz_qty_sold = oz_qty_returned = 0


# ── Расчёт P&L ────────────────────────────────────────────────────────────────
total_payout = wb_net + oz_net

tax_nds  = total_payout * nds_pct  / 100
tax_usn  = total_payout * usn_pct  / 100
tax_ins  = total_payout * ins_pct  / 100
tax_total= tax_nds + tax_usn + tax_ins

net_income = total_payout - tax_total


# ── Шапка: KPI ────────────────────────────────────────────────────────────────
st.markdown(f"### {calendar.month_name[month]} {year}")

k1, k2, k3, k4 = st.columns(4)
k1.metric("Выплата WB",   f"{wb_net:,.0f} ₽".replace(",", " "),
          help="ppvz_for_pay (продажи + возвраты) − логистика − хранение − штрафы")
k2.metric("Выплата Ozon", f"{oz_net:,.0f} ₽".replace(",", " "),
          help="Сумма всех транзакций финансового API Ozon (продажи + возвраты + прочие удержания)")
k3.metric("Итого выплат", f"{total_payout:,.0f} ₽".replace(",", " "))
k4.metric(f"Налоги {total_tax_pct:.0f}%", f"−{tax_total:,.0f} ₽".replace(",", " "),
          delta=f"{net_income:,.0f} ₽ чистыми".replace(",", " "),
          delta_color="normal")

st.markdown("---")


# ── P&L таблица ───────────────────────────────────────────────────────────────
st.markdown("#### Отчёт о прибылях и убытках")

def fmt(v: float, sign: bool = False) -> str:
    s_ = "+" if sign and v > 0 else ("−" if v < 0 else "")
    return f"{s_}{abs(v):,.0f} ₽".replace(",", " ")

rows = []

# Доходы — WB
rows.append({"Статья": "🏪 **Wildberries**",                           "Сумма (₽)": "",                     "Тип": "header"})
rows.append({"Статья": "   Выплата за продажи (ppvz_for_pay)",          "Сумма (₽)": fmt(wb_ppvz_sales),    "Тип": "detail"})
if abs(wb_ppvz_returns) > 0:
    rows.append({"Статья": "   Возвраты (ppvz_for_pay)",                "Сумма (₽)": fmt(wb_ppvz_returns),  "Тип": "detail"})
rows.append({"Статья": "   − Логистика (доставка/возвраты)",            "Сумма (₽)": fmt(-wb_delivery),     "Тип": "detail"})
rows.append({"Статья": "   − Хранение",                                 "Сумма (₽)": fmt(-wb_storage),      "Тип": "detail"})
if wb_penalties > 0:
    rows.append({"Статья": "   − Штрафы",                              "Сумма (₽)": fmt(-wb_penalties),    "Тип": "detail"})
if abs(wb_other) > 0:
    rows.append({"Статья": "   ± Прочее",                              "Сумма (₽)": fmt(wb_other),         "Тип": "detail"})
rows.append({"Статья": "**Итого выплата WB**",                          "Сумма (₽)": fmt(wb_net),           "Тип": "subtotal"})

rows.append({"Статья": "", "Сумма (₽)": "", "Тип": "spacer"})

# Доходы — Ozon
rows.append({"Статья": "🟦 **Ozon**",                                   "Сумма (₽)": "",                     "Тип": "header"})
rows.append({"Статья": "   Выплата за доставленные товары",             "Сумма (₽)": fmt(oz_sales_amount),   "Тип": "detail"})
if oz_returns_amount != 0:
    rows.append({"Статья": "   Возвраты и невыкупы",                    "Сумма (₽)": fmt(oz_returns_amount), "Тип": "detail"})
if oz_other_amount != 0:
    rows.append({"Статья": "   Прочие удержания (хранение, эквайринг)", "Сумма (₽)": fmt(oz_other_amount),   "Тип": "detail"})
rows.append({"Статья": "**Итого выплата Ozon**",                        "Сумма (₽)": fmt(oz_net),            "Тип": "subtotal"})

rows.append({"Статья": "", "Сумма (₽)": "", "Тип": "spacer"})

# Итого выплат
rows.append({"Статья": "**💰 ИТОГО ВЫПЛАТ от маркетплейсов**", "Сумма (₽)": fmt(total_payout), "Тип": "total"})

rows.append({"Статья": "", "Сумма (₽)": "", "Тип": "spacer"})

# Налоги
rows.append({"Статья": "🏛️ **Налоговая нагрузка**",     "Сумма (₽)": "",           "Тип": "header"})
rows.append({"Статья": f"   НДС {nds_pct:.1f}%",         "Сумма (₽)": fmt(-tax_nds),  "Тип": "tax"})
rows.append({"Статья": f"   УСН {usn_pct:.1f}%",          "Сумма (₽)": fmt(-tax_usn),  "Тип": "tax"})
rows.append({"Статья": f"   Страховые взносы {ins_pct:.1f}%", "Сумма (₽)": fmt(-tax_ins), "Тип": "tax"})
rows.append({"Статья": f"**Итого налоги {total_tax_pct:.1f}%**", "Сумма (₽)": fmt(-tax_total), "Тип": "tax_total"})

rows.append({"Статья": "", "Сумма (₽)": "", "Тип": "spacer"})

# Чистая выручка
rows.append({"Статья": "✅ **ЧИСТАЯ ВЫРУЧКА**",          "Сумма (₽)": fmt(net_income), "Тип": "net"})

pnl_df = pd.DataFrame(rows)[["Статья", "Сумма (₽)"]]
st.dataframe(
    pnl_df,
    use_container_width=True,
    hide_index=True,
    height=min(50 + len(pnl_df) * 35, 750),
)

st.markdown("---")


# ── Waterfall: от выплат к чистой выручке ─────────────────────────────────────
st.markdown("#### Структура: выплаты → чистая выручка")

labels = ["Выплата WB", "Выплата Ozon", "Итого выплат", f"НДС {nds_pct:.0f}%",
          f"УСН {usn_pct:.0f}%", f"Взносы {ins_pct:.0f}%", "Чистая выручка"]
measure = ["absolute", "relative", "total", "relative", "relative", "relative", "total"]
values  = [wb_net, oz_net, total_payout, -tax_nds, -tax_usn, -tax_ins, net_income]

fig = go.Figure(go.Waterfall(
    orientation="v",
    measure=measure,
    x=labels,
    y=values,
    text=[f"{abs(v):,.0f}".replace(",", " ") for v in values],
    textposition="outside",
    connector={"line": {"color": "#9CA3AF"}},
    increasing={"marker": {"color": "#10B981"}},
    decreasing={"marker": {"color": "#EF4444"}},
    totals={"marker": {"color": "#7C3AED"}},
))
fig.update_layout(margin=dict(t=40, b=10), height=400, showlegend=False)
st.plotly_chart(fig, use_container_width=True)

st.markdown("---")


# ── Доля налогов vs чистая выручка ────────────────────────────────────────────
col_pie, col_stats = st.columns([1, 1])

with col_pie:
    st.markdown("#### Распределение выплат")
    fig_pie = go.Figure(go.Pie(
        labels=[f"НДС {nds_pct:.0f}%", f"УСН {usn_pct:.0f}%",
                f"Взносы {ins_pct:.0f}%", "Чистая выручка"],
        values=[tax_nds, tax_usn, tax_ins, max(net_income, 0)],
        hole=0.45,
        marker_colors=["#EF4444", "#F59E0B", "#FCD34D", "#10B981"],
        textinfo="label+percent",
    ))
    fig_pie.update_layout(margin=dict(t=10, b=10), height=300, showlegend=False)
    st.plotly_chart(fig_pie, use_container_width=True)

with col_stats:
    st.markdown("#### Ключевые показатели")
    stats = {
        "Продано WB (шт.)":          f"{wb_qty_sold:,}".replace(",", " "),
        "Возвратов WB (шт.)":        f"{wb_qty_returned:,}".replace(",", " "),
        "Продано Ozon (шт.)":        f"{oz_qty_sold:,}".replace(",", " "),
        "Возвратов Ozon (шт.)":      f"{oz_qty_returned:,}".replace(",", " "),
        "Итого продано":             f"{wb_qty_sold + oz_qty_sold:,}".replace(",", " "),
        "Средняя выплата WB / шт":   (f"{wb_net/wb_qty_sold:,.0f} ₽".replace(",", " ")
                                       if wb_qty_sold > 0 else "—"),
        "Средняя выплата Ozon / шт": (f"{oz_sales_amount/oz_qty_sold:,.0f} ₽".replace(",", " ")
                                       if oz_qty_sold > 0 else "—"),
        "Налоговая нагрузка":        f"{tax_total/total_payout*100:.1f}%" if total_payout else "—",
        "Чистая выручка":            f"{net_income:,.0f} ₽".replace(",", " "),
    }
    for k, v in stats.items():
        st.markdown(f"**{k}:** {v}")


st.markdown("---")


# ── Excel экспорт ──────────────────────────────────────────────────────────────
def build_excel() -> bytes:
    wb_xl = Workbook()
    ws = wb_xl.active
    ws.title = f"PnL {calendar.month_abbr[month]}{year}"

    orange = PatternFill("solid", fgColor="F97316")
    light  = PatternFill("solid", fgColor="FED7AA")
    red_f  = PatternFill("solid", fgColor="FEE2E2")
    green_f= PatternFill("solid", fgColor="DCFCE7")
    gray_f = PatternFill("solid", fgColor="F3F4F6")
    bold   = Font(bold=True)
    thin   = Side(style="thin")

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 20

    # Заголовок
    ws.merge_cells("A1:B1")
    c = ws["A1"]
    c.value = f"P&L отчёт — {calendar.month_name[month]} {year}"
    c.font  = Font(bold=True, size=14)
    c.fill  = orange
    c.font  = Font(bold=True, size=14, color="FFFFFF")
    c.alignment = Alignment(horizontal="center")

    row = 3
    type_style = {
        "header":   (bold, gray_f),
        "subtotal": (Font(bold=True), light),
        "total":    (Font(bold=True, size=12), PatternFill("solid", fgColor="7C3AED")),
        "tax":      (Font(italic=True), red_f),
        "tax_total":(Font(bold=True), red_f),
        "net":      (Font(bold=True, size=12), green_f),
        "detail":   (Font(), None),
        "spacer":   (Font(), None),
    }

    for _, r in pd.DataFrame(rows).iterrows():
        t = r["Тип"]
        label_raw = r["Статья"].replace("**", "").replace("*", "")
        val_raw   = r["Сумма (₽)"]

        ca = ws.cell(row, 1, label_raw)
        cb = ws.cell(row, 2, val_raw)

        font_, fill_ = type_style.get(t, (Font(), None))
        ca.font = font_
        cb.font = font_
        cb.alignment = Alignment(horizontal="right")

        if fill_:
            ca.fill = fill_
            cb.fill = fill_

        if t == "total":
            ca.font = Font(bold=True, size=12, color="FFFFFF")
            cb.font = Font(bold=True, size=12, color="FFFFFF")

        row += 1

    buf = io.BytesIO()
    wb_xl.save(buf)
    return buf.getvalue()


st.download_button(
    "⬇️ Скачать Excel (P&L)",
    data=build_excel(),
    file_name=f"pnl_{year}_{month:02d}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

"""
Расчёт налоговой базы и налоговой нагрузки ИП (УСН + НДС).

Методология ФНС (агентская схема):
  • Налоговая база = цена продавца ДО скидок маркетплейса (retail_price × qty)
  • Общий оборот   = фактическая реализация покупателям (retail_price_withdisc_rub × qty)
  • НДС 5%  — «изнутри» базы: база × 5/105
  • УСН 6%  — на (база − НДС) × 6%
"""

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import calendar
import io
from datetime import date as date_cls
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.data_loader import load_wb_financial, load_ozon_postings
from src.auth import require_role

require_role(["admin"])
st.title("🧾 Расчёт налога")

if "database" not in st.secrets:
    st.error("Настройте базу данных в разделе ⚙️ Настройки.")
    st.stop()

DB_URL = st.secrets["database"]["url"]

# ── Период ────────────────────────────────────────────────────────────────────
today = date_cls.today()
col_y, col_m = st.sidebar.columns(2)
year  = col_y.selectbox("Год",   list(range(today.year, today.year - 3, -1)), index=0)
month = col_m.selectbox("Месяц", list(range(1, 13)), index=today.month - 1,
                        format_func=lambda m: calendar.month_name[m])
st.sidebar.markdown("---")

# ── Налоговые ставки ──────────────────────────────────────────────────────────
st.sidebar.markdown("### Ставки налогов")
nds_pct = st.sidebar.number_input("НДС (%)",     value=5.0, min_value=0.0, max_value=30.0, step=0.5)
usn_pct = st.sidebar.number_input("УСН (%)",     value=6.0, min_value=0.0, max_value=15.0, step=0.5)
ins_pct = st.sidebar.number_input("1% СФР (%)", value=1.0, min_value=0.0, max_value=5.0, step=0.1)
st.sidebar.caption(
    f"НДС считается **изнутри** базы: база × {nds_pct:.0f}/{100+nds_pct:.0f}\n\n"
    f"УСН применяется к (база − НДС)"
)
st.sidebar.markdown("---")
st.sidebar.caption(f"Период: {calendar.month_name[month]} {year}")


# ── Загрузка данных ───────────────────────────────────────────────────────────
df_wb = load_wb_financial(DB_URL, year, month)
df_oz = load_ozon_postings(DB_URL, year, month)


# ── WB: оборот и база ─────────────────────────────────────────────────────────
wb_empty = df_wb.empty
wb_oborot = wb_base = 0.0
wb_qty_sold = wb_qty_returned = 0
wb_sales_ob = wb_ret_ob = wb_sales_base = wb_ret_base = 0.0

if not wb_empty:
    sales_mask   = df_wb["doc_type_name"] == "Продажа"
    returns_mask = df_wb["doc_type_name"] == "Возврат"

    df_sales = df_wb[sales_mask].copy()
    df_ret   = df_wb[returns_mask].copy()

    df_sales["qty"] = df_sales["quantity"].fillna(0).clip(lower=0)
    df_ret["qty"]   = df_ret["quantity"].fillna(0).abs()

    # Общий оборот = фактическая реализация (цена с учётом скидки маркетплейса)
    wb_sales_ob  = float((df_sales["retail_price_withdisc_rub"].fillna(0) * df_sales["qty"]).sum())
    wb_ret_ob    = float((df_ret["retail_price_withdisc_rub"].fillna(0)   * df_ret["qty"]).sum())
    wb_oborot    = wb_sales_ob - wb_ret_ob

    # Налоговая база ФНС = цена продавца до скидок
    wb_sales_base = float((df_sales["retail_price"].fillna(0) * df_sales["qty"]).sum())
    wb_ret_base   = float((df_ret["retail_price"].fillna(0)   * df_ret["qty"]).sum())
    wb_base       = wb_sales_base - wb_ret_base

    wb_qty_sold     = int(df_sales["qty"].sum())
    wb_qty_returned = int(df_ret["qty"].sum())


# ── Ozon: оборот и база ───────────────────────────────────────────────────────
oz_empty = df_oz.empty
oz_oborot = oz_base = 0.0
oz_qty_sold = 0

if not oz_empty:
    if "is_cancelled" in df_oz.columns:
        df_active = df_oz[df_oz["is_cancelled"] != True].copy()
    else:
        df_active = df_oz.copy()

    df_active["qty"] = df_active["quantity"].fillna(0).clip(lower=0)

    # Налоговая база Ozon = payout (реализация) — это «Доход» в финансовом дашборде Ozon.
    # price × qty давало завышенную базу (~10M), payout соответствует «Отчёту о реализации» (~5.9M).
    if "payout" in df_active.columns:
        oz_base = float(df_active["payout"].fillna(0).sum())
    else:
        oz_base = float((df_active["price"].fillna(0) * df_active["qty"]).sum())
    oz_oborot  = oz_base
    oz_qty_sold = int(df_active["qty"].sum())


# ── Налоговая нагрузка (от базы ФНС) ─────────────────────────────────────────
total_oborot = wb_oborot + oz_oborot
total_base   = wb_base   + oz_base

nds_rate = nds_pct / (100 + nds_pct)
tax_nds  = total_base * nds_rate
tax_usn  = (total_base - tax_nds) * usn_pct / 100
tax_ins  = max(0.0, (total_base - 300_000) * ins_pct / 100) if ins_pct > 0 else 0.0
tax_total = tax_nds + tax_usn + tax_ins


# ── KPI карточки ──────────────────────────────────────────────────────────────
st.markdown(f"### {calendar.month_name[month]} {year}")

k1, k2, k3, k4 = st.columns(4)
k1.metric(
    "Общий оборот",
    f"{total_oborot:,.0f} ₽".replace(",", " "),
    help="Фактическая реализация покупателям (цена с учётом скидок маркетплейса)"
)
k2.metric(
    "Налоговая база ФНС",
    f"{total_base:,.0f} ₽".replace(",", " "),
    help="Цена продавца ДО скидок маркетплейса — позиция ФНС по агентской схеме"
)
k3.metric(
    f"НДС {nds_pct:.0f}%",
    f"{tax_nds:,.0f} ₽".replace(",", " "),
    help=f"НДС изнутри: база × {nds_pct:.0f}/{100+nds_pct:.0f}"
)
k4.metric(
    f"УСН {usn_pct:.0f}%",
    f"{tax_usn:,.0f} ₽".replace(",", " "),
    help=f"(база − НДС) × {usn_pct:.0f}%"
)

st.markdown("---")

k5, k6, k7, k8 = st.columns(4)
k5.metric("Продано WB (шт.)",   wb_qty_sold)
k6.metric("Возвратов WB (шт.)", wb_qty_returned)
k7.metric("Продано Ozon (шт.)", oz_qty_sold)
k8.metric(
    "Итого налоговая нагрузка",
    f"{tax_total:,.0f} ₽".replace(",", " "),
    delta=f"{total_base - tax_total:,.0f} ₽ остаток от базы".replace(",", " "),
    delta_color="normal",
)

st.markdown("---")


# ── Таблица расчёта ───────────────────────────────────────────────────────────
st.markdown("#### Расчёт налоговой базы")


def fmt(v: float) -> str:
    if v == 0:
        return "—"
    sign = "−" if v < 0 else ""
    return f"{sign}{abs(v):,.0f} ₽".replace(",", " ")


rows_tbl = []

# WB
rows_tbl.append({"Статья": "🏪 Wildberries", "Оборот (₽)": "", "База ФНС (₽)": "", "Тип": "header"})
rows_tbl.append({
    "Статья":       f"   Продажи (шт. {wb_qty_sold})",
    "Оборот (₽)":   fmt(wb_sales_ob),
    "База ФНС (₽)": fmt(wb_sales_base),
    "Тип": "detail"
})
if wb_qty_returned > 0:
    rows_tbl.append({
        "Статья":       f"   − Возвраты (шт. {wb_qty_returned})",
        "Оборот (₽)":   fmt(-wb_ret_ob),
        "База ФНС (₽)": fmt(-wb_ret_base),
        "Тип": "detail"
    })
rows_tbl.append({
    "Статья":       "Итого WB",
    "Оборот (₽)":   fmt(wb_oborot),
    "База ФНС (₽)": fmt(wb_base),
    "Тип": "subtotal"
})

rows_tbl.append({"Статья": "", "Оборот (₽)": "", "База ФНС (₽)": "", "Тип": "spacer"})

# Ozon
rows_tbl.append({"Статья": "🟦 Ozon", "Оборот (₽)": "", "База ФНС (₽)": "", "Тип": "header"})
rows_tbl.append({
    "Статья":       f"   Продажи (шт. {oz_qty_sold})",
    "Оборот (₽)":   fmt(oz_oborot),
    "База ФНС (₽)": fmt(oz_base),
    "Тип": "detail"
})
rows_tbl.append({
    "Статья":       "Итого Ozon",
    "Оборот (₽)":   fmt(oz_oborot),
    "База ФНС (₽)": fmt(oz_base),
    "Тип": "subtotal"
})

rows_tbl.append({"Статья": "", "Оборот (₽)": "", "База ФНС (₽)": "", "Тип": "spacer"})

rows_tbl.append({
    "Статья":       "💰 ИТОГО",
    "Оборот (₽)":   fmt(total_oborot),
    "База ФНС (₽)": fmt(total_base),
    "Тип": "total"
})

rows_tbl.append({"Статья": "", "Оборот (₽)": "", "База ФНС (₽)": "", "Тип": "spacer"})

rows_tbl.append({"Статья": "🏛️ Налоговая нагрузка (от базы ФНС)", "Оборот (₽)": "", "База ФНС (₽)": "", "Тип": "header"})
rows_tbl.append({
    "Статья":       f"   НДС {nds_pct:.1f}% — изнутри базы",
    "Оборот (₽)":   "",
    "База ФНС (₽)": fmt(-tax_nds),
    "Тип": "tax"
})
rows_tbl.append({
    "Статья":       f"   УСН {usn_pct:.1f}% × (база − НДС)",
    "Оборот (₽)":   "",
    "База ФНС (₽)": fmt(-tax_usn),
    "Тип": "tax"
})
if ins_pct > 0 and tax_ins > 0:
    rows_tbl.append({
        "Статья":       "   1% СФР на сверхдоход",
        "Оборот (₽)":   "",
        "База ФНС (₽)": fmt(-tax_ins),
        "Тип": "tax"
    })
rows_tbl.append({
    "Статья":       "Итого налоги",
    "Оборот (₽)":   "",
    "База ФНС (₽)": fmt(-tax_total),
    "Тип": "tax_total"
})

rows_tbl.append({"Статья": "", "Оборот (₽)": "", "База ФНС (₽)": "", "Тип": "spacer"})

rows_tbl.append({
    "Статья":       "✅ База после налогов",
    "Оборот (₽)":   "",
    "База ФНС (₽)": fmt(total_base - tax_total),
    "Тип": "net"
})

pnl_df = pd.DataFrame(rows_tbl)[["Статья", "Оборот (₽)", "База ФНС (₽)"]]
st.dataframe(
    pnl_df,
    use_container_width=True,
    hide_index=True,
    height=min(60 + len(pnl_df) * 35, 800),
)

st.markdown("---")


# ── Waterfall ─────────────────────────────────────────────────────────────────
st.markdown("#### Налоговая база → после налогов")

labels_wf  = ["База ФНС", f"НДС {nds_pct:.0f}%", f"УСН {usn_pct:.0f}%"]
measure_wf = ["absolute", "relative", "relative"]
values_wf  = [total_base, -tax_nds, -tax_usn]

if ins_pct > 0 and tax_ins > 0:
    labels_wf.append("1% СФР")
    measure_wf.append("relative")
    values_wf.append(-tax_ins)

labels_wf.append("После налогов")
measure_wf.append("total")
values_wf.append(total_base - tax_total)

fig = go.Figure(go.Waterfall(
    orientation="v",
    measure=measure_wf,
    x=labels_wf,
    y=values_wf,
    text=[f"{abs(v):,.0f}".replace(",", " ") for v in values_wf],
    textposition="outside",
    connector={"line": {"color": "#9CA3AF"}},
    increasing={"marker": {"color": "#10B981"}},
    decreasing={"marker": {"color": "#EF4444"}},
    totals={"marker": {"color": "#7C3AED"}},
))
fig.update_layout(margin=dict(t=40, b=10), height=380, showlegend=False)
st.plotly_chart(fig, use_container_width=True)

st.markdown("---")


# ── Donut + таблица ───────────────────────────────────────────────────────────
col_pie, col_stats = st.columns([1, 1])

with col_pie:
    st.markdown("#### Распределение налоговой базы")
    pie_labels = [f"НДС {nds_pct:.0f}%", f"УСН {usn_pct:.0f}%"]
    pie_values = [tax_nds, tax_usn]
    pie_colors = ["#EF4444", "#F59E0B"]
    if ins_pct > 0 and tax_ins > 0:
        pie_labels.append("1% СФР")
        pie_values.append(tax_ins)
        pie_colors.append("#3B82F6")
    pie_labels.append("После налогов")
    pie_values.append(max(total_base - tax_total, 0))
    pie_colors.append("#10B981")

    fig_pie = go.Figure(go.Pie(
        labels=pie_labels,
        values=[round(v) for v in pie_values],
        hole=0.45,
        marker_colors=pie_colors,
        textinfo="label+percent",
        hovertemplate="%{label}<br>%{value:,.0f} ₽<extra></extra>",
    ))
    fig_pie.update_layout(margin=dict(t=10, b=10), height=300, showlegend=False)
    st.plotly_chart(fig_pie, use_container_width=True)

with col_stats:
    st.markdown("#### Итоговые показатели")
    eff_rate = tax_total / total_base * 100 if total_base else 0
    diff_ob  = total_base - total_oborot
    net_after = total_base - tax_total

    st.markdown(f"""
| Показатель | Значение |
|---|---|
| Общий оборот | **{total_oborot:,.0f} ₽** |
| Налоговая база ФНС | **{total_base:,.0f} ₽** |
| Разница (скидки МП) | {diff_ob:,.0f} ₽ |
| НДС {nds_pct:.0f}% | {tax_nds:,.0f} ₽ |
| УСН {usn_pct:.0f}% | {tax_usn:,.0f} ₽ |
| Итого налоги | **{tax_total:,.0f} ₽** |
| База после налогов | **{net_after:,.0f} ₽** |
| Эффективная ставка | **{eff_rate:.1f}%** |
| Продано всего (шт.) | {wb_qty_sold + oz_qty_sold:,} |
""".replace(",", " "))

st.markdown("---")


# ── Excel-экспорт ─────────────────────────────────────────────────────────────
def build_excel() -> bytes:
    wb_xl = Workbook()
    ws = wb_xl.active
    ws.title = f"Налог {calendar.month_abbr[month]}{year}"

    purple = PatternFill("solid", fgColor="4C1D95")
    light  = PatternFill("solid", fgColor="EDE9FE")
    red_f  = PatternFill("solid", fgColor="FEE2E2")
    green_f= PatternFill("solid", fgColor="DCFCE7")
    gray_f = PatternFill("solid", fgColor="F3F4F6")

    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22

    ws.merge_cells("A1:C1")
    c = ws["A1"]
    c.value = f"Расчёт налога — {calendar.month_name[month]} {year}"
    c.font  = Font(bold=True, size=14, color="FFFFFF")
    c.fill  = purple
    c.alignment = Alignment(horizontal="center")

    for col_i, hdr in enumerate(["Статья", "Оборот (₽)", "База ФНС (₽)"], start=1):
        cell = ws.cell(row=2, column=col_i, value=hdr)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = purple
        cell.alignment = Alignment(horizontal="center" if col_i > 1 else "left")

    type_style = {
        "header":   (Font(bold=True),          gray_f),
        "subtotal": (Font(bold=True),           light),
        "total":    (Font(bold=True, size=12, color="FFFFFF"), purple),
        "tax":      (Font(italic=True),         red_f),
        "tax_total":(Font(bold=True),           red_f),
        "net":      (Font(bold=True),           green_f),
        "detail":   (Font(),                   None),
        "spacer":   (Font(),                   None),
    }

    row = 3
    for _, r in pd.DataFrame(rows_tbl).iterrows():
        t     = r["Тип"]
        label = str(r["Статья"])
        ob    = str(r["Оборот (₽)"])
        base  = str(r["База ФНС (₽)"])

        font_, fill_ = type_style.get(t, (Font(), None))
        for col_i, val in enumerate([label, ob, base], start=1):
            cell = ws.cell(row=row, column=col_i, value=val)
            cell.font = font_
            cell.alignment = Alignment(
                horizontal="right" if col_i > 1 else "left",
                vertical="center"
            )
            if fill_:
                cell.fill = fill_
        row += 1

    buf = io.BytesIO()
    wb_xl.save(buf)
    return buf.getvalue()


st.download_button(
    "⬇️ Скачать Excel",
    data=build_excel(),
    file_name=f"расчёт_налога_{year}_{month:02d}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

"""
Налоговая нагрузка ИП (УСН 6% + НДС 5%).

Методология:
  • База ФНС = цена продавца до скидок маркетплейса (ФНС: агент не меняет базу).
    - WB:   retail_price × quantity  (строки «Продажа»)
    - Ozon: price × quantity         (из ozon_postings, статус не «cancelled»)
  • Фактически к перечислению на р/с:
    - WB:   ppvz_for_pay (все строки, включая услуги/хранение)
    - Ozon: amount (все OperationAgentDeliveredToCustomer-операции)
  • НДС 5% — «изнутри» суммы: база × 5/105
  • УСН 6%  — на (база − НДС): (база − НДС) × 6%
  • 1% СФР  — 1% × MAX(0, годовая_база − 300 000), cap = 300 888 ₽
  • Фиксированные взносы ИП 2026: 53 658 ₽/год (показываются отдельно)
"""

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
import calendar
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

from src.db.models import init_db, get_session_factory
from src.db.repository import FinancialReportRepository, OzonPostingRepository, OzonTransactionRepository
from src.auth import require_role

require_role(["admin"])
st.title("🧾 Налоговая нагрузка ИП")

if "database" not in st.secrets:
    st.error("Настройте базу данных в разделе ⚙️ Настройки.")
    st.stop()

DB_URL = st.secrets["database"]["url"]

# ── Константы ─────────────────────────────────────────────────────────────────
NDS_RATE       = 5 / 105      # НДС «изнутри» суммы при ставке 5%
USN_RATE       = 0.06         # УСН доходы 6%
SFR_RATE       = 0.01         # 1% взнос СФР сверх 300 000
SFR_THRESHOLD  = 300_000      # порог для 1% СФР
SFR_CAP        = 300_888      # максимум 1% СФР в 2026 году
IP_FIXED       = 53_658       # фиксированные взносы ИП 2026

MONTH_NAMES = {
    1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель",
    5: "Май", 6: "Июнь", 7: "Июль", 8: "Август",
    9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь",
}

# ── Выбор периода ─────────────────────────────────────────────────────────────
from datetime import date as date_cls
today = date_cls.today()

year = st.sidebar.selectbox("Год", list(range(today.year, today.year - 3, -1)), index=0)
st.sidebar.markdown("---")
st.sidebar.caption(
    "База ФНС = цена продавца (до скидок маркетплейса).\n\n"
    "НДС 5% считается «изнутри» базы: база × 5/105.\n\n"
    "УСН 6% — на (база − НДС)."
)

# ── Загрузка данных ───────────────────────────────────────────────────────────
@st.cache_data(ttl=300, show_spinner="Загружаю данные WB...")
def load_wb(db_url: str, year: int, month: int) -> pd.DataFrame:
    engine = init_db(db_url)
    Session = get_session_factory(engine)
    with Session() as session:
        return FinancialReportRepository().get_by_month(session, year, month)


@st.cache_data(ttl=300, show_spinner="Загружаю данные Ozon...")
def load_ozon_postings(db_url: str, year: int, month: int) -> pd.DataFrame:
    engine = init_db(db_url)
    Session = get_session_factory(engine)
    with Session() as session:
        return OzonPostingRepository().get_by_month(session, year, month)


@st.cache_data(ttl=300, show_spinner="Загружаю транзакции Ozon...")
def load_ozon_tx(db_url: str, year: int, month: int) -> pd.DataFrame:
    engine = init_db(db_url)
    Session = get_session_factory(engine)
    with Session() as session:
        return OzonTransactionRepository().get_by_month(session, year, month)


def calc_wb_month(year: int, month: int) -> dict:
    df = load_wb(DB_URL, year, month)
    if df.empty:
        return dict(base=0, payout=0, has_data=False)

    sales = df[df["doc_type_name"] == "Продажа"].copy()
    ret   = df[df["doc_type_name"] == "Возврат"].copy()

    # База ФНС = retail_price × quantity (продажи) − retail_price × |quantity| (возвраты)
    sales["qty"] = sales["quantity"].fillna(0).clip(lower=0)
    ret["qty"]   = ret["quantity"].fillna(0).abs()

    base_sales = (sales["retail_price"].fillna(0) * sales["qty"]).sum()
    base_ret   = (ret["retail_price"].fillna(0) * ret["qty"]).sum()
    base = base_sales - base_ret

    # К перечислению = ppvz_for_pay по всем строкам (включая услуги, хранение, штрафы)
    payout = float(df["ppvz_for_pay"].fillna(0).sum())

    return dict(base=float(base), payout=float(payout), has_data=True)


def calc_ozon_month(year: int, month: int) -> dict:
    postings = load_ozon_postings(DB_URL, year, month)
    tx       = load_ozon_tx(DB_URL, year, month)

    # База ФНС из постингов: price × quantity, только не отменённые
    base = 0.0
    if not postings.empty:
        p = postings[postings.get("is_cancelled", pd.Series(False, index=postings.index)) != True].copy()
        p["qty"] = p["quantity"].fillna(0).clip(lower=0)
        base = float((p["price"].fillna(0) * p["qty"]).sum())

    # К перечислению = сумма amount по всем транзакциям месяца
    payout = 0.0
    if not tx.empty:
        payout = float(tx["amount"].fillna(0).sum())

    has_data = (not postings.empty) or (not tx.empty)
    return dict(base=float(base), payout=float(payout), has_data=has_data)


def calc_taxes(base_wb: float, base_ozon: float) -> dict:
    base_total = base_wb + base_ozon
    nds        = base_total * NDS_RATE
    usn_base   = base_total - nds
    usn        = usn_base * USN_RATE
    total_tax  = nds + usn
    return dict(
        base_total=base_total,
        nds=nds,
        usn=usn,
        total_tax=total_tax,
    )


# ── Расчёт по всем месяцам ────────────────────────────────────────────────────
rows = []
with st.spinner("Считаю налоговую нагрузку по месяцам..."):
    for m in range(1, 13):
        wb   = calc_wb_month(year, m)
        oz   = calc_ozon_month(year, m)
        if not wb["has_data"] and not oz["has_data"]:
            continue
        taxes = calc_taxes(wb["base"], oz["base"])
        rows.append({
            "month":      m,
            "month_name": MONTH_NAMES[m],
            "base_wb":    wb["base"],
            "base_ozon":  oz["base"],
            "base_total": taxes["base_total"],
            "nds":        taxes["nds"],
            "usn":        taxes["usn"],
            "total_tax":  taxes["total_tax"],
            "payout_wb":  wb["payout"],
            "payout_ozon":oz["payout"],
        })

if not rows:
    st.warning(f"Нет данных за {year} год. Загрузите финансовые отчёты WB и Ozon.")
    st.stop()

df_months = pd.DataFrame(rows)

# ── Итоги за год ──────────────────────────────────────────────────────────────
tot_base_wb    = df_months["base_wb"].sum()
tot_base_ozon  = df_months["base_ozon"].sum()
tot_base       = df_months["base_total"].sum()
tot_nds        = df_months["nds"].sum()
tot_usn        = df_months["usn"].sum()
tot_tax        = df_months["total_tax"].sum()
tot_payout_wb  = df_months["payout_wb"].sum()
tot_payout_ozon= df_months["payout_ozon"].sum()
tot_payout     = tot_payout_wb + tot_payout_ozon

# 1% СФР — считается на годовую базу
sfr_1pct = min(max(0.0, (tot_base - SFR_THRESHOLD) * SFR_RATE), float(SFR_CAP))
total_ip_burden = tot_tax + sfr_1pct + IP_FIXED   # всё включая фиксированные взносы

# ── KPI карточки ──────────────────────────────────────────────────────────────
st.markdown(f"### Итоги за {year} год")

c1, c2, c3, c4 = st.columns(4)
c1.metric(
    "База ФНС (цена продавца)",
    f"{tot_base:,.0f} ₽".replace(",", " "),
    help="Сумма цен продавца до скидок маркетплейса — налогооблагаемая база по позиции ФНС"
)
c2.metric(
    "НДС 5%",
    f"{tot_nds:,.0f} ₽".replace(",", " "),
    help="НДС «изнутри» базы: база × 5 / 105"
)
c3.metric(
    "УСН 6%",
    f"{tot_usn:,.0f} ₽".replace(",", " "),
    help="УСН на (база − НДС) × 6%"
)
c4.metric(
    "Итого налоговая нагрузка",
    f"{tot_tax:,.0f} ₽".replace(",", " "),
    help="НДС + УСН"
)

st.markdown("---")

c5, c6, c7, c8 = st.columns(4)
c5.metric(
    "Фактически получено на р/с",
    f"{tot_payout:,.0f} ₽".replace(",", " "),
    help="WB: ppvz_for_pay (к перечислению). Ozon: amount всех транзакций."
)
c6.metric(
    "1% СФР (сверх 300 тыс.)",
    f"{sfr_1pct:,.0f} ₽".replace(",", " "),
    help=f"1% × (база − 300 000), максимум {SFR_CAP:,} ₽".replace(",", " ")
)
c7.metric(
    "Фиксированные взносы ИП",
    f"{IP_FIXED:,} ₽".replace(",", " "),
    help="Фиксированные взносы ИП 2026 (53 658 ₽/год)"
)
c8.metric(
    "Всего обязательства ИП",
    f"{total_ip_burden:,.0f} ₽".replace(",", " "),
    help="НДС + УСН + 1% СФР + фиксированные взносы",
    delta=f"{tot_payout - total_ip_burden:,.0f} ₽ остаток".replace(",", " "),
    delta_color="normal" if tot_payout > total_ip_burden else "inverse",
)

st.markdown("---")

# ── Помесячная таблица ────────────────────────────────────────────────────────
st.markdown("#### Помесячная разбивка")

display_df = df_months[[
    "month_name", "base_wb", "base_ozon", "base_total",
    "nds", "usn", "total_tax", "payout_wb", "payout_ozon"
]].copy()

# Строка итогов
totals_row = pd.DataFrame([{
    "month_name": "ИТОГО",
    "base_wb":    tot_base_wb,
    "base_ozon":  tot_base_ozon,
    "base_total": tot_base,
    "nds":        tot_nds,
    "usn":        tot_usn,
    "total_tax":  tot_tax,
    "payout_wb":  tot_payout_wb,
    "payout_ozon":tot_payout_ozon,
}])
display_df = pd.concat([display_df, totals_row], ignore_index=True)

st.dataframe(
    display_df.rename(columns={
        "month_name":  "Месяц",
        "base_wb":     "База ФНС WB (₽)",
        "base_ozon":   "База ФНС Ozon (₽)",
        "base_total":  "База ФНС Итого (₽)",
        "nds":         "НДС 5% (₽)",
        "usn":         "УСН 6% (₽)",
        "total_tax":   "Налог Итого (₽)",
        "payout_wb":   "К перечислению WB (₽)",
        "payout_ozon": "К перечислению Ozon (₽)",
    }),
    use_container_width=True,
    hide_index=True,
    column_config={
        "База ФНС WB (₽)":        st.column_config.NumberColumn(format="%.0f"),
        "База ФНС Ozon (₽)":      st.column_config.NumberColumn(format="%.0f"),
        "База ФНС Итого (₽)":     st.column_config.NumberColumn(format="%.0f"),
        "НДС 5% (₽)":             st.column_config.NumberColumn(format="%.0f"),
        "УСН 6% (₽)":             st.column_config.NumberColumn(format="%.0f"),
        "Налог Итого (₽)":        st.column_config.NumberColumn(format="%.0f"),
        "К перечислению WB (₽)":  st.column_config.NumberColumn(format="%.0f"),
        "К перечислению Ozon (₽)":st.column_config.NumberColumn(format="%.0f"),
    },
)

st.markdown("---")

# ── График: налоговая нагрузка по месяцам ────────────────────────────────────
st.markdown("#### Налоговая нагрузка по месяцам")

fig = go.Figure()
fig.add_bar(
    name="НДС 5%",
    x=df_months["month_name"],
    y=df_months["nds"].round(0),
    marker_color="#EF4444",
    text=df_months["nds"].apply(lambda v: f"{v:,.0f}".replace(",", " ")),
    textposition="inside",
)
fig.add_bar(
    name="УСН 6%",
    x=df_months["month_name"],
    y=df_months["usn"].round(0),
    marker_color="#F59E0B",
    text=df_months["usn"].apply(lambda v: f"{v:,.0f}".replace(",", " ")),
    textposition="inside",
)
fig.update_layout(
    barmode="stack",
    height=360,
    margin=dict(t=20, b=10),
    legend=dict(orientation="h", y=1.05),
    yaxis_title="Сумма (₽)",
)
st.plotly_chart(fig, use_container_width=True)

# ── График: база vs реализация ───────────────────────────────────────────────
st.markdown("#### База ФНС vs Фактически на р/с по месяцам")

fig2 = go.Figure()
fig2.add_bar(
    name="База ФНС WB",
    x=df_months["month_name"],
    y=df_months["base_wb"].round(0),
    marker_color="#7C3AED",
)
fig2.add_bar(
    name="База ФНС Ozon",
    x=df_months["month_name"],
    y=df_months["base_ozon"].round(0),
    marker_color="#A855F7",
)
fig2.add_scatter(
    name="К перечислению WB",
    x=df_months["month_name"],
    y=df_months["payout_wb"].round(0),
    mode="lines+markers",
    line=dict(color="#10B981", width=2),
    marker=dict(size=6),
)
fig2.add_scatter(
    name="К перечислению Ozon",
    x=df_months["month_name"],
    y=df_months["payout_ozon"].round(0),
    mode="lines+markers",
    line=dict(color="#06B6D4", width=2, dash="dot"),
    marker=dict(size=6),
)
fig2.update_layout(
    barmode="group",
    height=380,
    margin=dict(t=20, b=10),
    legend=dict(orientation="h", y=1.08),
    yaxis_title="Сумма (₽)",
)
st.plotly_chart(fig2, use_container_width=True)

st.markdown("---")

# ── Структура годовой налоговой нагрузки (доnut) ─────────────────────────────
st.markdown("#### Структура налоговой нагрузки за год")

col_pie, col_info = st.columns([1, 1])

with col_pie:
    pie_labels = ["НДС 5%", "УСН 6%", "1% СФР", "Фикс. взносы ИП"]
    pie_values = [tot_nds, tot_usn, sfr_1pct, IP_FIXED]
    pie_colors = ["#EF4444", "#F59E0B", "#3B82F6", "#8B5CF6"]
    fig_pie = go.Figure(go.Pie(
        labels=pie_labels,
        values=[round(v) for v in pie_values],
        hole=0.45,
        marker_colors=pie_colors,
        textinfo="label+percent",
        hovertemplate="%{label}<br>%{value:,.0f} ₽<extra></extra>",
    ))
    fig_pie.update_layout(
        height=320,
        margin=dict(t=10, b=10, l=0, r=0),
        showlegend=False,
    )
    st.plotly_chart(fig_pie, use_container_width=True)

with col_info:
    eff_rate_base  = (tot_tax / tot_base * 100) if tot_base else 0
    eff_rate_payout= (total_ip_burden / tot_payout * 100) if tot_payout else 0
    net_after_tax  = tot_payout - total_ip_burden

    st.markdown("**Ключевые показатели**")
    st.markdown(f"""
| Показатель | Сумма |
|---|---|
| База ФНС (цена продавца) | **{tot_base:,.0f} ₽** |
| НДС 5% | {tot_nds:,.0f} ₽ |
| УСН 6% | {tot_usn:,.0f} ₽ |
| НДС + УСН | **{tot_tax:,.0f} ₽** |
| 1% СФР | {sfr_1pct:,.0f} ₽ |
| Фиксированные взносы ИП | {IP_FIXED:,} ₽ |
| **Всего обязательства ИП** | **{total_ip_burden:,.0f} ₽** |
| Фактически на р/с | {tot_payout:,.0f} ₽ |
| Осталось после налогов | **{net_after_tax:,.0f} ₽** |
| Эффективная ставка (налог/база) | {eff_rate_base:.1f}% |
| Нагрузка на р/с | **{eff_rate_payout:.1f}%** |
""".replace(",", " "))

st.markdown("---")

# ── Экспорт в Excel ───────────────────────────────────────────────────────────
st.markdown("#### Экспорт")

def build_excel() -> bytes:
    wb_xl = Workbook()
    ws = wb_xl.active
    ws.title = f"Налоги {year}"

    # Стили
    hdr_font  = Font(bold=True, color="FFFFFF", size=10)
    hdr_fill  = PatternFill("solid", start_color="4C1D95")
    tot_fill  = PatternFill("solid", start_color="EDE9FE")
    tot_font  = Font(bold=True, size=10)
    center    = Alignment(horizontal="center", vertical="center")
    num_fmt   = '#,##0'
    thin      = Side(style="thin", color="D1D5DB")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "Месяц",
        "База ФНС WB (₽)", "База ФНС Ozon (₽)", "База ФНС Итого (₽)",
        "НДС 5% (₽)", "УСН 6% (₽)", "Налог Итого (₽)",
        "К перечислению WB (₽)", "К перечислению Ozon (₽)",
    ]
    col_widths = [14, 20, 22, 22, 16, 16, 18, 22, 24]

    # Заголовок
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = f"Налоговая нагрузка ИП за {year} год"
    title_cell.font = Font(bold=True, size=13, color="4C1D95")
    title_cell.alignment = center

    ws.row_dimensions[1].height = 24

    # Строка заголовков
    for col_i, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=2, column=col_i, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col_i)].width = w

    ws.row_dimensions[2].height = 30

    # Данные
    data_rows = df_months[df_months["month_name"] != "ИТОГО"]
    for row_i, row in enumerate(data_rows.itertuples(), start=3):
        values = [
            row.month_name,
            row.base_wb, row.base_ozon, row.base_total,
            row.nds, row.usn, row.total_tax,
            row.payout_wb, row.payout_ozon,
        ]
        for col_i, val in enumerate(values, start=1):
            cell = ws.cell(row=row_i, column=col_i, value=round(val) if isinstance(val, float) else val)
            cell.border = border
            cell.alignment = Alignment(horizontal="right" if col_i > 1 else "left", vertical="center")
            if col_i > 1:
                cell.number_format = num_fmt
        ws.row_dimensions[row_i].height = 18

    # Итоговая строка
    tot_row = 3 + len(data_rows)
    tot_values = [
        "ИТОГО",
        tot_base_wb, tot_base_ozon, tot_base,
        tot_nds, tot_usn, tot_tax,
        tot_payout_wb, tot_payout_ozon,
    ]
    for col_i, val in enumerate(tot_values, start=1):
        cell = ws.cell(row=tot_row, column=col_i, value=round(val) if isinstance(val, float) else val)
        cell.font = tot_font
        cell.fill = tot_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="right" if col_i > 1 else "left", vertical="center")
        if col_i > 1:
            cell.number_format = num_fmt
    ws.row_dimensions[tot_row].height = 20

    # Блок взносов ИП
    extra_start = tot_row + 2
    ip_rows = [
        ("1% СФР (на сверхдоход)", sfr_1pct),
        ("Фиксированные взносы ИП", IP_FIXED),
        ("Всего обязательства ИП (НДС+УСН+СФР+фикс.)", total_ip_burden),
        ("Осталось на р/с после налогов", tot_payout - total_ip_burden),
    ]
    for i, (label, val) in enumerate(ip_rows):
        r = extra_start + i
        ws.merge_cells(f"A{r}:G{r}")
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = Font(italic=True, size=10)
        vc = ws.cell(row=r, column=8, value=round(val))
        vc.number_format = num_fmt
        vc.font = Font(bold=(i == 2 or i == 3), size=10)

    buf = io.BytesIO()
    wb_xl.save(buf)
    buf.seek(0)
    return buf.read()


excel_bytes = build_excel()
st.download_button(
    label="📥 Скачать Excel",
    data=excel_bytes,
    file_name=f"налоги_ИП_{year}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

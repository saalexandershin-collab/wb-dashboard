"""
Сверка: начислено маркетплейсом vs. фактически на р/с.

Три уровня данных:
  1. Оборот (реализация) — что покупатели заплатили маркетплейсу
  2. Начислено МП        — ppvz_for_pay WB / amount Ozon (после всех удержаний)
  3. Фактически на р/с  — реальные банковские поступления из выписки

Дебиторка МП = накопленное начислено − накопленное на р/с
"""

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import calendar
import io
from datetime import date as date_cls
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from src.data_loader import load_wb_financial_range, load_ozon_postings_range, load_ozon_transactions_range
from src.auth import require_role

require_role(["admin"])
st.title("🔍 Сверка: начислено МП vs. поступило на р/с")

if "database" not in st.secrets:
    st.error("Настройте базу данных в разделе ⚙️ Настройки.")
    st.stop()

DB_URL = st.secrets["database"]["url"]

# ── Фактические поступления на р/с (из банковских выписок) ───────────────────
# Обновляй этот словарь при каждом новом поступлении.
# Ключ: (год, месяц) — месяц, когда деньги ПРИШЛИ на р/с.

WB_BANK: dict[tuple, float] = {
    # 2025
    (2025,  1):  5_643_557,
    (2025,  2):  4_013_794,
    (2025,  3):  4_186_341,
    (2025,  4):  5_677_586,
    (2025,  5):  6_283_157,
    (2025,  6): 10_049_072,
    (2025,  7):  9_151_644,
    (2025,  8):  3_778_611,
    (2025,  9):  7_039_030,
    (2025, 10):  4_962_261,
    (2025, 11):  9_265_856,
    (2025, 12):  9_308_687,
    # 2026
    (2026,  1):  7_446_715,
    (2026,  2): 10_050_889,
    (2026,  3):  4_702_393,
    (2026,  4):  7_395_070,
    (2026,  5):  4_887_776,
}

OZON_BANK: dict[tuple, float] = {
    # 2026 (в файле подписано как 2025, но фактически 2026 год)
    (2026,  1):   368_052,
    (2026,  2): 1_451_509,
    (2026,  3): 1_099_038,
    (2026,  4): 1_333_822,
    (2026,  5):   436_876,
}

# ── Период ──────────────────────────────────────────────────────────────────
today = date_cls.today()
st.sidebar.markdown("### Период сверки")
col1, col2 = st.sidebar.columns(2)
year_from  = col1.selectbox("С года",   list(range(today.year, today.year - 3, -1)), index=0, key="yf")
month_from = col2.selectbox("С месяца", list(range(1, 13)), index=0, key="mf",
                             format_func=lambda m: calendar.month_abbr[m])
col3, col4 = st.sidebar.columns(2)
year_to  = col3.selectbox("По год",    list(range(today.year, today.year - 3, -1)), index=0, key="yt")
month_to = col4.selectbox("По месяц",  list(range(1, 13)), index=today.month - 1, key="mt",
                           format_func=lambda m: calendar.month_abbr[m])
st.sidebar.markdown("---")
st.sidebar.caption(
    "**Начислено МП** — ppvz_for_pay WB и amount Ozon (из финансовых отчётов).\n\n"
    "**На р/с** — фактические банковские поступления из выписки (обновляются вручную в коде).\n\n"
    "**Дебиторка** = накопленное начислено − накопленное на р/с."
)

# ── Загрузка данных из БД ────────────────────────────────────────────────────
wb_raw = load_wb_financial_range(DB_URL, year_from, month_from, year_to, month_to)
oz_raw = load_ozon_postings_range(DB_URL, year_from, month_from, year_to, month_to)
oz_tx  = load_ozon_transactions_range(DB_URL, year_from, month_from, year_to, month_to)

# ── Расчёт WB (оборот + начислено МП) ────────────────────────────────────────
def calc_wb(df):
    if df.empty:
        return pd.DataFrame()
    rows = []
    for (y, m), g in df.groupby(["_year", "_month"]):
        sales   = g[g["doc_type_name"] == "Продажа"]
        returns = g[g["doc_type_name"] == "Возврат"]

        oborot_net = (
            (sales["retail_price_withdisc_rub"] * sales["quantity"].abs()).sum()
            - (returns["retail_price_withdisc_rub"] * returns["quantity"].abs()).sum()
        )
        ppvz_net   = g["ppvz_for_pay"].sum()

        commission = g["ppvz_sales_commission"].fillna(0).sum()
        delivery   = g["delivery_rub"].fillna(0).sum()
        storage    = g["storage_fee"].fillna(0).sum()
        penalty    = g["penalty"].fillna(0).sum()
        acquiring  = g["acquiring_fee"].fillna(0).sum()
        total_fees = abs(commission) + abs(delivery) + abs(storage) + abs(penalty) + abs(acquiring)

        rows.append({
            "year": y, "month": m,
            "wb_oborot": oborot_net,
            "wb_ppvz":   ppvz_net,
            "wb_commission": abs(commission),
            "wb_delivery":   abs(delivery),
            "wb_storage":    abs(storage),
            "wb_penalty":    abs(penalty),
            "wb_acquiring":  abs(acquiring),
            "wb_fees":       total_fees,
        })
    return pd.DataFrame(rows).sort_values(["year", "month"])


# ── Расчёт Ozon (оборот + начислено МП) ──────────────────────────────────────
def calc_ozon(postings, transactions):
    posting_rows, tx_map = [], {}
    if not postings.empty:
        for (y, m), g in postings.groupby(["_year", "_month"]):
            active = g[~g["is_cancelled"]] if "is_cancelled" in g.columns else g
            oborot = (
                active["payout"].fillna(0).sum() if "payout" in active.columns
                else (active["price"].fillna(0) * active["quantity"].abs()).sum()
            )
            posting_rows.append({"year": y, "month": m, "oz_oborot": oborot})
    if not transactions.empty and "amount" in transactions.columns:
        for (y, m), g in transactions.groupby(["_year", "_month"]):
            tx_map[(y, m)] = g["amount"].sum()

    months_set = {(r["year"], r["month"]) for r in posting_rows} | set(tx_map)
    result = []
    for (y, m) in sorted(months_set):
        oz_oborot = next((r["oz_oborot"] for r in posting_rows if r["year"] == y and r["month"] == m), 0)
        result.append({"year": y, "month": m, "oz_oborot": oz_oborot, "oz_ppvz": tx_map.get((y, m), 0)})
    return pd.DataFrame(result).sort_values(["year", "month"]) if result else pd.DataFrame()


df_wb = calc_wb(wb_raw)
df_oz = calc_ozon(oz_raw, oz_tx)

# ── Сборка сводной таблицы ───────────────────────────────────────────────────
months_all = sorted(set(
    (list(zip(df_wb["year"], df_wb["month"])) if not df_wb.empty else []) +
    (list(zip(df_oz["year"], df_oz["month"])) if not df_oz.empty else [])
))

# Добавляем месяцы, у которых есть только банковские поступления (без данных в БД)
bank_months = set(WB_BANK) | set(OZON_BANK)
for ym in bank_months:
    if year_from <= ym[0] <= year_to:
        if (ym[0], month_from) <= ym <= (ym[0], month_to) or ym[0] in range(year_from, year_to + 1):
            if ym not in months_all:
                months_all.append(ym)
months_all = sorted(set(months_all))

if not months_all:
    st.warning("Нет данных за выбранный период.")
    st.stop()

# Фильтруем по выбранному периоду
months_all = [
    (y, m) for (y, m) in months_all
    if (year_from, month_from) <= (y, m) <= (year_to, month_to)
]

summary_rows = []
for (y, m) in months_all:
    wb_r  = df_wb[(df_wb["year"] == y) & (df_wb["month"] == m)]
    oz_r  = df_oz[(df_oz["year"] == y) & (df_oz["month"] == m)] if not df_oz.empty else pd.DataFrame()

    wb_oborot = float(wb_r["wb_oborot"].iloc[0]) if not wb_r.empty else 0.0
    wb_ppvz   = float(wb_r["wb_ppvz"].iloc[0])   if not wb_r.empty else 0.0
    wb_fees   = float(wb_r["wb_fees"].iloc[0])    if not wb_r.empty else 0.0
    oz_oborot = float(oz_r["oz_oborot"].iloc[0])  if not oz_r.empty else 0.0
    oz_ppvz   = float(oz_r["oz_ppvz"].iloc[0])    if not oz_r.empty else 0.0

    wb_bank   = WB_BANK.get((y, m), 0.0)
    oz_bank   = OZON_BANK.get((y, m), 0.0)

    total_oborot  = wb_oborot + oz_oborot
    total_ppvz    = wb_ppvz + oz_ppvz
    total_bank    = wb_bank + oz_bank

    summary_rows.append({
        "Период":           f"{calendar.month_name[m]} {y}",
        "_y": y, "_m": m,
        "WB оборот":        wb_oborot,
        "WB начислено МП":  wb_ppvz,
        "WB на р/с":        wb_bank,
        "Ozon оборот":      oz_oborot,
        "Ozon начислено МП":oz_ppvz,
        "Ozon на р/с":      oz_bank,
        "Итого оборот":     total_oborot,
        "Итого начислено":  total_ppvz,
        "Итого на р/с":     total_bank,
        "WB удержания":     wb_fees,
    })

df_sum = pd.DataFrame(summary_rows)

# Накопительная дебиторка: накопленное начислено − накопленное на р/с
df_sum["_cum_ppvz"] = df_sum["Итого начислено"].cumsum()
df_sum["_cum_bank"] = df_sum["Итого на р/с"].cumsum()
df_sum["Дебиторка МП"] = df_sum["_cum_ppvz"] - df_sum["_cum_bank"]

# ── KPI карточки ─────────────────────────────────────────────────────────────
tot_oborot  = df_sum["Итого оборот"].sum()
tot_ppvz    = df_sum["Итого начислено"].sum()
tot_bank    = df_sum["Итого на р/с"].sum()
tot_wb_fees = df_sum["WB удержания"].sum()
debitorka   = df_sum["Дебиторка МП"].iloc[-1]  # текущая дебиторка (накопленная)

st.markdown("### Итого за период")
c1, c2, c3, c4, c5 = st.columns(5)
c1.metric("Оборот (реализация)", f"{tot_oborot:,.0f} ₽".replace(",", " "),
          help="Что покупатели заплатили маркетплейсам")
c2.metric("Начислено МП",        f"{tot_ppvz:,.0f} ₽".replace(",", " "),
          help="ppvz_for_pay WB + amount Ozon (после комиссий и удержаний)")
c3.metric("Фактически на р/с",   f"{tot_bank:,.0f} ₽".replace(",", " "),
          help="Реальные банковские поступления из выписки")
c4.metric("Разница (начислено − р/с)", f"{tot_ppvz - tot_bank:,.0f} ₽".replace(",", " "),
          help="Положительное = МП начислило больше, чем пришло на р/с (ожидаемые поступления)")
c5.metric("Дебиторка МП (накопл.)", f"{debitorka:,.0f} ₽".replace(",", " "),
          delta=f"{'▲' if debitorka > 0 else '▼'} задолженность МП перед тобой" if debitorka != 0 else "сверка закрыта",
          delta_color="off",
          help="Накопленное начислено − накопленное на р/с. "
               "Положительное = деньги у МП, ещё не перечислены.")

st.markdown("---")

# ── Основная таблица сверки ───────────────────────────────────────────────────
st.subheader("📋 Сверка по месяцам")
st.caption(
    "**Начислено МП** — из финансовых отчётов (ppvz_for_pay WB, amount Ozon). "
    "**На р/с** — фактические банковские поступления. "
    "**Дебиторка** — накопленная разница (нарастающим итогом)."
)

disp_cols = [
    "Период",
    "WB начислено МП", "WB на р/с",
    "Ozon начислено МП", "Ozon на р/с",
    "Итого начислено", "Итого на р/с",
    "Дебиторка МП",
]
disp = df_sum[disp_cols].copy()

# Строка итогов (дебиторка — последнее значение накопленного, не сумма)
totals_row = {
    "Период": "ИТОГО",
    "WB начислено МП":   df_sum["WB начислено МП"].sum(),
    "WB на р/с":         df_sum["WB на р/с"].sum(),
    "Ozon начислено МП": df_sum["Ozon начислено МП"].sum(),
    "Ozon на р/с":       df_sum["Ozon на р/с"].sum(),
    "Итого начислено":   df_sum["Итого начислено"].sum(),
    "Итого на р/с":      df_sum["Итого на р/с"].sum(),
    "Дебиторка МП":      debitorka,   # накопленная, не сумма строк
}
disp = pd.concat([disp, pd.DataFrame([totals_row])], ignore_index=True)

money_cols = [c for c in disp.columns if c != "Период"]

def style_row(row):
    styles = []
    for col in disp.columns:
        if row["Период"] == "ИТОГО":
            styles.append("font-weight:bold; background-color:#EDE9FE")
        elif col == "Дебиторка МП":
            val = row[col]
            if isinstance(val, (int, float)):
                styles.append("color:#EF4444" if val > 0 else "color:#10B981")
            else:
                styles.append("")
        else:
            styles.append("")
    return styles

st.dataframe(
    disp.style
        .format({c: lambda v: f"{v:,.0f} ₽".replace(",", " ") if isinstance(v, (int, float)) else v
                 for c in money_cols})
        .apply(style_row, axis=1),
    use_container_width=True,
    hide_index=True,
    height=min(80 + len(disp) * 38, 600),
)

st.markdown("---")

# ── График: начислено МП vs фактически на р/с ────────────────────────────────
st.subheader("📊 Начислено МП vs. фактически поступило на р/с")

labels = df_sum["Период"].tolist()
fig1 = go.Figure()
fig1.add_bar(
    name="Начислено МП (WB)", x=labels,
    y=df_sum["WB начислено МП"].round(0),
    marker_color="#7C3AED",
)
fig1.add_bar(
    name="Начислено МП (Ozon)", x=labels,
    y=df_sum["Ozon начислено МП"].round(0),
    marker_color="#A855F7",
)
fig1.add_scatter(
    name="Фактически на р/с (WB)", x=labels,
    y=df_sum["WB на р/с"].round(0),
    mode="lines+markers",
    line=dict(color="#10B981", width=2),
    marker=dict(size=7),
)
fig1.add_scatter(
    name="Фактически на р/с (Ozon)", x=labels,
    y=df_sum["Ozon на р/с"].round(0),
    mode="lines+markers",
    line=dict(color="#06B6D4", width=2, dash="dot"),
    marker=dict(size=7),
)
fig1.update_layout(
    barmode="stack", height=380,
    margin=dict(t=20, b=10),
    yaxis_title="₽",
    legend=dict(orientation="h", y=1.1),
)
st.plotly_chart(fig1, use_container_width=True)

# ── График: накопительная дебиторка ──────────────────────────────────────────
st.subheader("📈 Накопительная дебиторка МП (нарастающим итогом)")

fig2 = go.Figure()
fig2.add_scatter(
    name="Накоплено начислено МП", x=labels,
    y=df_sum["_cum_ppvz"].round(0),
    mode="lines+markers",
    line=dict(color="#7C3AED", width=2),
    fill="tozeroy", fillcolor="rgba(124,58,237,0.07)",
)
fig2.add_scatter(
    name="Накоплено на р/с", x=labels,
    y=df_sum["_cum_bank"].round(0),
    mode="lines+markers",
    line=dict(color="#10B981", width=2),
    fill="tozeroy", fillcolor="rgba(16,185,129,0.07)",
)
fig2.add_scatter(
    name="Дебиторка МП", x=labels,
    y=df_sum["Дебиторка МП"].round(0),
    mode="lines+markers",
    line=dict(color="#EF4444", width=2, dash="dot"),
    marker=dict(size=7, symbol="diamond"),
)
fig2.update_layout(
    height=340,
    margin=dict(t=20, b=10),
    yaxis_title="₽ (нарастающим итогом)",
    legend=dict(orientation="h", y=1.1),
)
st.plotly_chart(fig2, use_container_width=True)

st.markdown("---")

# ── Детализация удержаний WB ─────────────────────────────────────────────────
if not df_wb.empty:
    with st.expander("🔎 Детализация удержаний WB по месяцам"):
        wb_det = df_wb[["year", "month", "wb_commission", "wb_delivery",
                         "wb_storage", "wb_penalty", "wb_acquiring", "wb_fees"]].copy()
        wb_det["Период"] = wb_det.apply(lambda r: f"{calendar.month_name[int(r['month'])]} {int(r['year'])}", axis=1)
        wb_det = wb_det[["Период", "wb_commission", "wb_delivery", "wb_storage",
                          "wb_penalty", "wb_acquiring", "wb_fees"]]
        wb_det.columns = ["Период", "Комиссия WB", "Логистика", "Хранение", "Штрафы", "Эквайринг", "Итого удержания"]

        tot_det = {"Период": "ИТОГО", **{c: wb_det[c].sum() for c in wb_det.columns if c != "Период"}}
        wb_det = pd.concat([wb_det, pd.DataFrame([tot_det])], ignore_index=True)
        st.dataframe(
            wb_det.style
                  .format({c: lambda v: f"{v:,.0f} ₽".replace(",", " ") if isinstance(v, (int, float)) else v
                           for c in wb_det.columns if c != "Период"})
                  .apply(lambda row: ["font-weight:bold" if row["Период"] == "ИТОГО" else "" for _ in row], axis=1),
            use_container_width=True, hide_index=True,
        )

# ── Пояснение ────────────────────────────────────────────────────────────────
st.info(
    "**Как читать дебиторку:**\n\n"
    "- **Положительная дебиторка** — МП начислило тебе больше, чем пришло на р/с. "
    "Это нормально: выплаты идут с задержкой 7–14 дней (WB) или 2 раза в месяц (Ozon).\n\n"
    "- **Отрицательная** — на р/с пришло больше, чем начислено в выбранном периоде. "
    "Типично для января: декабрьские начисления WB приходят в январе.\n\n"
    "- Данные 'На р/с' обновляются **вручную** в файле `pages/13_🔍_Сверка.py` "
    "в словарях `WB_BANK` и `OZON_BANK`.",
    icon="ℹ️"
)

# ── Excel-экспорт ────────────────────────────────────────────────────────────
st.markdown("---")
def build_excel() -> bytes:
    wb_xls = Workbook()
    ws = wb_xls.active
    ws.title = "Сверка"

    hdr_font   = Font(bold=True, color="FFFFFF")
    hdr_fill   = PatternFill("solid", start_color="4C1D95")
    total_fill = PatternFill("solid", start_color="EDE9FE")
    red_fill   = PatternFill("solid", start_color="FEE2E2")
    green_fill = PatternFill("solid", start_color="DCFCE7")
    center     = Alignment(horizontal="center")
    money_fmt  = '#,##0'

    headers = [
        "Период",
        "WB начислено МП", "WB на р/с",
        "Ozon начислено МП", "Ozon на р/с",
        "Итого начислено", "Итого на р/с",
        "Дебиторка МП (накопл.)",
    ]
    col_w = [20, 20, 16, 22, 16, 20, 16, 24]

    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    ws["A1"].value = f"Сверка: начислено МП vs. поступило на р/с  |  период {calendar.month_abbr[month_from]} {year_from} – {calendar.month_abbr[month_to]} {year_to}"
    ws["A1"].font  = Font(bold=True, size=12, color="4C1D95")
    ws["A1"].alignment = center

    for ci, (h, w) in enumerate(zip(headers, col_w), 1):
        c = ws.cell(2, ci, h)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = center
        ws.column_dimensions[get_column_letter(ci)].width = w

    for ri, row in enumerate(summary_rows, 3):
        y_, m_ = row["_y"], row["_m"]
        vals = [
            row["Период"],
            row["WB начислено МП"], WB_BANK.get((y_, m_), 0),
            row["Ozon начислено МП"], OZON_BANK.get((y_, m_), 0),
            row["Итого начислено"], row["Итого на р/с"],
            df_sum[df_sum["_m"] == m_]["Дебиторка МП"].values[0] if m_ in df_sum["_m"].values else 0,
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(ri, ci, round(v) if isinstance(v, float) else v)
            if ci > 1:
                cell.number_format = money_fmt
                cell.alignment = Alignment(horizontal="right")
            if ci == len(headers) and isinstance(v, (int, float)):
                cell.fill = red_fill if v > 0 else green_fill

    # Итог
    tr = len(summary_rows) + 3
    tot_vals = [
        "ИТОГО",
        df_sum["WB начислено МП"].sum(),   df_sum["WB на р/с"].sum(),
        df_sum["Ozon начислено МП"].sum(), df_sum["Ozon на р/с"].sum(),
        df_sum["Итого начислено"].sum(),   df_sum["Итого на р/с"].sum(),
        debitorka,
    ]
    for ci, v in enumerate(tot_vals, 1):
        cell = ws.cell(tr, ci, round(v) if isinstance(v, float) else v)
        cell.font = Font(bold=True); cell.fill = total_fill
        if ci > 1:
            cell.number_format = money_fmt
            cell.alignment = Alignment(horizontal="right")

    buf = io.BytesIO()
    wb_xls.save(buf)
    buf.seek(0)
    return buf.read()

if st.button("📥 Скачать сверку Excel"):
    st.download_button(
        "⬇️ Скачать сверку.xlsx",
        build_excel(),
        f"сверка_{year_from}_{month_from:02d}–{year_to}_{month_to:02d}.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

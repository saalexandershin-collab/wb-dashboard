import streamlit as st
import pandas as pd
import calendar
import io
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter

from src.data_loader import load_wb_orders, load_wb_sales
st.title("📋 Детализация по товарам")

if "database" not in st.secrets or "wildberries" not in st.secrets:
    st.error("Сначала настройте токен и базу данных в разделе ⚙️ Настройки.")
    st.stop()

DB_URL = st.secrets["database"]["url"]

# ── Период ───────────────────────────────────────────────────────────────────
today = date.today()
col_y, col_m = st.sidebar.columns(2)
year = col_y.selectbox("Год", list(range(today.year, today.year - 3, -1)))
month = col_m.selectbox("Месяц", list(range(1, 13)), index=today.month - 1,
                        format_func=lambda m: calendar.month_name[m])

# ── Загрузка ─────────────────────────────────────────────────────────────────
df_orders = load_wb_orders(DB_URL, year, month)
df_sales  = load_wb_sales(DB_URL, year, month)

if df_orders.empty and df_sales.empty:
    st.warning("Нет данных за выбранный период.")
    st.stop()

# ── Подготовка ────────────────────────────────────────────────────────────────
def prep_orders(df):
    if df.empty:
        return df
    df = df.copy()
    df["order_date"] = pd.to_datetime(df["order_date"])
    df["day"] = df["order_date"].dt.day
    df["is_actual"] = ~df["is_cancel"].fillna(False)
    return df

def prep_sales(df):
    if df.empty:
        return df
    df = df.copy()
    df["sale_date"] = pd.to_datetime(df["sale_date"])
    df["day"] = df["sale_date"].dt.day
    return df

ord_df = prep_orders(df_orders)
sal_df = prep_sales(df_sales)

# ── Фильтры в боковом меню ───────────────────────────────────────────────────
st.sidebar.markdown("---")
st.sidebar.markdown("### Фильтры")

def sidebar_filter(label, series):
    vals = sorted(series.dropna().unique().tolist())
    return st.sidebar.multiselect(label, vals)

brands = sidebar_filter("Бренд", ord_df["brand"] if not ord_df.empty else pd.Series(dtype=str))
articles = sidebar_filter("Артикул продавца", ord_df["supplier_article"] if not ord_df.empty else pd.Series(dtype=str))
warehouses = sidebar_filter("Склад", ord_df["warehouse_name"] if not ord_df.empty else pd.Series(dtype=str))
regions = sidebar_filter("Регион", ord_df["region_name"] if not ord_df.empty else pd.Series(dtype=str))

def apply_filters(df, col_map):
    for col, vals in col_map.items():
        if vals and col in df.columns:
            df = df[df[col].isin(vals)]
    return df

filter_map = {
    "brand": brands,
    "supplier_article": articles,
    "warehouse_name": warehouses,
    "region_name": regions,
}
ord_df = apply_filters(ord_df, filter_map)
sal_df = apply_filters(sal_df, filter_map)

# ── Переключатель показателя ─────────────────────────────────────────────────
st.markdown("### Показатель")
metric = st.radio(
    "Выберите показатель для таблицы:",
    options=[
        "Заказы, шт.",
        "Выкупы, шт.",
        "Возвраты, шт.",
        "% выкупа",
    ],
    horizontal=True,
)

# ── Построение сводной таблицы ────────────────────────────────────────────────
days_in_month = calendar.monthrange(year, month)[1]
all_days = list(range(1, days_in_month + 1))

def build_pivot(df, day_col, value_col, agg="count", item_col="nm_id"):
    if df.empty:
        return pd.DataFrame()
    if agg == "count":
        piv = df.groupby([item_col, day_col]).size().reset_index(name="val")
    else:
        piv = df.groupby([item_col, day_col])[value_col].sum().reset_index(name="val")
    piv = piv.pivot(index=item_col, columns=day_col, values="val").reindex(columns=all_days).fillna(0)
    return piv

def build_info(df):
    if df.empty:
        return pd.DataFrame(columns=["nm_id", "brand", "supplier_article"])
    return df.groupby("nm_id").agg(
        brand=("brand", "last"),
        supplier_article=("supplier_article", "last"),
    ).reset_index()

# Строим нужную таблицу
if metric == "Заказы, шт.":
    actual = ord_df[ord_df["is_actual"]] if not ord_df.empty else ord_df
    piv = build_pivot(actual, "day", None, "count")
elif metric == "Выкупы, шт.":
    sub = sal_df[sal_df["is_return"] == False] if not sal_df.empty else sal_df
    piv = build_pivot(sub, "day", None, "count")
elif metric == "Возвраты, шт.":
    sub = sal_df[sal_df["is_return"] == True] if not sal_df.empty else sal_df
    piv = build_pivot(sub, "day", None, "count")
elif metric == "% выкупа":
    actual = ord_df[ord_df["is_actual"]] if not ord_df.empty else ord_df
    sub = sal_df[sal_df["is_return"] == False] if not sal_df.empty else sal_df
    o_piv = build_pivot(actual, "day", None, "count")
    b_piv = build_pivot(sub, "day", None, "count")
    all_idx = o_piv.index.union(b_piv.index)
    o_piv = o_piv.reindex(all_idx, fill_value=0)
    b_piv = b_piv.reindex(all_idx, fill_value=0)
    piv = (b_piv / o_piv.replace(0, float("nan")) * 100).round(1).fillna(0)
else:
    piv = pd.DataFrame()

if piv.empty:
    st.info("Нет данных для построения таблицы.")
    st.stop()

# ── Добавляем инфо о товаре и итоги ──────────────────────────────────────────
source_df = ord_df if not ord_df.empty else sal_df
info = build_info(source_df)
piv = piv.reset_index()
piv = piv.merge(info, on="nm_id", how="left")

# Итог по строке
day_cols = [c for c in piv.columns if isinstance(c, int)]
piv["Итого"] = piv[day_cols].sum(axis=1).round(1)

# Переименуем колонки в формат "01", "02" ...
rename = {d: f"{d:02d}" for d in day_cols}
piv = piv.rename(columns=rename)
str_day_cols = [f"{d:02d}" for d in day_cols]

# Итоговая строка по дням
total_row = {"nm_id": "", "brand": "", "supplier_article": "ИТОГО"}
for c in str_day_cols:
    total_row[c] = piv[c].sum().round(1)
total_row["Итого"] = piv["Итого"].sum().round(1)
piv = pd.concat([piv, pd.DataFrame([total_row])], ignore_index=True)

# Порядок колонок
display_cols = ["brand", "supplier_article", "nm_id"] + str_day_cols + ["Итого"]
piv = piv[display_cols].rename(columns={
    "brand": "Бренд",
    "supplier_article": "Артикул",
    "nm_id": "nmId WB",
})

# ── Отображение ───────────────────────────────────────────────────────────────
st.markdown(f"**{metric}** — {calendar.month_name[month]} {year}")
st.dataframe(piv, use_container_width=True, height=500)

# ── Экспорт в Excel ───────────────────────────────────────────────────────────
def to_excel(df: pd.DataFrame, metric_name: str, year: int, month: int) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = f"{calendar.month_abbr[month]}{year}"

    header_fill = PatternFill("solid", fgColor="7C3AED")
    total_fill = PatternFill("solid", fgColor="E9D5FF")
    header_font = Font(bold=True, color="FFFFFF")
    total_font = Font(bold=True)

    ws.cell(1, 1, f"Показатель: {metric_name}").font = Font(bold=True, size=12)
    ws.cell(2, 1, f"Период: {calendar.month_name[month]} {year}").font = Font(italic=True)

    start_row = 4
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(start_row, col_idx, str(col_name))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        is_total = str(row[1]) == "ИТОГО"
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row_idx, col_idx, val)
            if is_total:
                cell.font = total_font
                cell.fill = total_fill
            if col_idx > 3:
                cell.number_format = "#,##0"

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 20)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

excel_data = to_excel(piv, metric, year, month)
st.download_button(
    label="⬇️ Скачать Excel",
    data=excel_data,
    file_name=f"wb_{year}_{month:02d}_{metric[:10].replace(' ', '_')}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

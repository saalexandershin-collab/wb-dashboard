import streamlit as st
import pandas as pd
import calendar
import io
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from src.data_loader import load_ozon_postings, load_ozon_transactions

st.title("📋 Детализация по товарам Ozon")

if "database" not in st.secrets:
    st.error("Настройте базу данных в разделе ⚙️ Настройки.")
    st.stop()

DB_URL = st.secrets["database"]["url"]

today = date.today()
col_y, col_m = st.sidebar.columns(2)
year  = col_y.selectbox("Год",   list(range(today.year, today.year - 3, -1)))
month = col_m.selectbox("Месяц", list(range(1, 13)), index=today.month - 1,
                        format_func=lambda m: calendar.month_name[m])

posts = load_ozon_postings(DB_URL, year, month)
txs   = load_ozon_transactions(DB_URL, year, month)

if posts.empty and txs.empty:
    st.warning("Нет данных за выбранный период.")
    st.stop()

if not posts.empty:
    posts["created_at"] = pd.to_datetime(posts["created_at"])
    posts["day"] = posts["created_at"].dt.day

if not txs.empty:
    txs["operation_date"] = pd.to_datetime(txs["operation_date"])
    txs["day"] = txs["operation_date"].dt.day

# ── Фильтры ───────────────────────────────────────────────────────────────────
st.sidebar.markdown("---")
st.sidebar.markdown("### Фильтры")

all_offers = sorted(set(
    (posts["offer_id"].dropna().unique().tolist() if not posts.empty else []) +
    (txs["offer_id"].dropna().unique().tolist() if not txs.empty else [])
))
sel_offers = st.sidebar.multiselect("Артикул (offer_id)", all_offers)

if sel_offers:
    if not posts.empty:
        posts = posts[posts["offer_id"].isin(sel_offers)]
    if not txs.empty:
        txs = txs[txs["offer_id"].isin(sel_offers)]

# ── Метрика ───────────────────────────────────────────────────────────────────
st.markdown("### Показатель")
metric = st.radio(
    "Выберите показатель:",
    ["Заказы (шт.)", "Выкупы (шт.)", "Отмены (шт.)"],
    horizontal=True,
)

days_in_month = calendar.monthrange(year, month)[1]
all_days = list(range(1, days_in_month + 1))

def build_pivot_posts(mask, qty_col="quantity"):
    sub = posts[mask] if not posts.empty else pd.DataFrame()
    if sub.empty:
        return pd.DataFrame()
    piv = (sub.groupby(["offer_id", "day"])[qty_col].sum()
             .reset_index()
             .pivot(index="offer_id", columns="day", values=qty_col)
             .reindex(columns=all_days).fillna(0))
    return piv

def build_pivot_tx():
    """Выкупы из транзакций — каждая строка = 1 единица."""
    sub = txs[txs["operation_type"] == "OperationAgentDeliveredToCustomer"] if not txs.empty else pd.DataFrame()
    if sub.empty:
        return pd.DataFrame()
    piv = (sub.groupby(["offer_id", "day"]).size()
             .reset_index(name="qty")
             .pivot(index="offer_id", columns="day", values="qty")
             .reindex(columns=all_days).fillna(0))
    return piv

if metric == "Заказы (шт.)":
    piv = build_pivot_posts(~posts["is_cancelled"]) if not posts.empty else pd.DataFrame()
    name_source = posts
elif metric == "Выкупы (шт.)":
    piv = build_pivot_tx()
    name_source = txs
else:
    piv = build_pivot_posts(posts["is_cancelled"]) if not posts.empty else pd.DataFrame()
    name_source = posts

if piv is None or piv.empty:
    st.info("Нет данных.")
    st.stop()

piv = piv.reset_index()

# Добавляем название товара
if not name_source.empty and "product_name" in name_source.columns:
    name_map = name_source.drop_duplicates("offer_id").set_index("offer_id")["product_name"]
    piv["product_name"] = piv["offer_id"].map(name_map).fillna("")
else:
    piv["product_name"] = ""

day_cols = [c for c in piv.columns if isinstance(c, int)]
piv["Итого"] = piv[day_cols].sum(axis=1).round(1)
rename = {d: f"{d:02d}" for d in day_cols}
piv = piv.rename(columns=rename)
str_day_cols = [f"{d:02d}" for d in day_cols]

total_row = {"offer_id": "", "product_name": "ИТОГО"}
for c in str_day_cols:
    total_row[c] = piv[c].sum().round(1)
total_row["Итого"] = piv["Итого"].sum().round(1)
piv = pd.concat([piv, pd.DataFrame([total_row])], ignore_index=True)

display_cols = ["offer_id", "product_name"] + str_day_cols + ["Итого"]
piv = piv[display_cols].rename(columns={"offer_id": "Артикул", "product_name": "Название"})

if metric == "Выкупы (шт.)":
    st.caption("Выкупы — по дате фактической доставки покупателю (транзакционная модель Ozon)")

st.markdown(f"**{metric}** — {calendar.month_name[month]} {year}")
st.dataframe(piv, use_container_width=True, height=500)

# ── Экспорт Excel ─────────────────────────────────────────────────────────────
def to_excel(df_: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = f"{calendar.month_abbr[month]}{year}"
    hfill = PatternFill("solid", fgColor="F97316")
    tfill = PatternFill("solid", fgColor="FED7AA")
    ws.cell(1, 1, f"Ozon — {metric}").font = Font(bold=True, size=12)
    ws.cell(2, 1, f"{calendar.month_name[month]} {year}").font = Font(italic=True)
    for ci, col in enumerate(df_.columns, 1):
        c = ws.cell(4, ci, str(col))
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = hfill
        c.alignment = Alignment(horizontal="center")
    for ri, row in enumerate(df_.itertuples(index=False), 5):
        is_total = str(row[0]) == "ИТОГО"
        for ci, val in enumerate(row, 1):
            cell = ws.cell(ri, ci, val)
            if is_total:
                cell.font = Font(bold=True)
                cell.fill = tfill
    for col in ws.columns:
        mx = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(mx + 2, 25)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

excel_data = to_excel(piv)
st.download_button(
    "⬇️ Скачать Excel",
    data=excel_data,
    file_name=f"ozon_{year}_{month:02d}_{metric[:10].replace(' ', '_')}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

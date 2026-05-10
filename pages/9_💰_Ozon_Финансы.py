import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import calendar
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.db.models import init_db, get_session_factory
from src.db.repository import OzonTransactionRepository
from src.auth import require_role

require_role(["admin"])
st.title("💰 Финансовый отчёт Ozon")

if "database" not in st.secrets:
    st.error("Настройте базу данных в разделе ⚙️ Настройки.")
    st.stop()

DB_URL = st.secrets["database"]["url"]

from datetime import date as date_cls
today = date_cls.today()
col_y, col_m = st.sidebar.columns(2)
year  = col_y.selectbox("Год",   list(range(today.year, today.year - 3, -1)), index=0)
month = col_m.selectbox("Месяц", list(range(1, 13)), index=today.month - 1,
                        format_func=lambda m: calendar.month_name[m])
st.sidebar.markdown("---")
st.sidebar.caption(f"Период: {calendar.month_name[month]} {year}")


@st.cache_data(ttl=300, show_spinner="Загружаю финансовые данные Ozon...")
def load_fin(db_url, year, month):
    engine = init_db(db_url)
    Session = get_session_factory(engine)
    with Session() as session:
        return OzonTransactionRepository().get_by_month(session, year, month)


df = load_fin(DB_URL, year, month)

if df.empty:
    st.warning("Нет финансовых данных Ozon за этот период.")
    st.code(
        f"OZON_CLIENT_ID='...' OZON_API_KEY='...' DATABASE_URL='...' "
        f"SYNC_YEAR={year} SYNC_MONTH={month} "
        "python3 scripts/sync_ozon_finances.py"
    )
    st.stop()

# ── Классификация операций ─────────────────────────────────────────────────────
# Ozon Finance API: каждая строка = одна операция.
# amount = чистая сумма зачисления (+) или удержания (-) на счёт продавца.
#
# Внутри OperationAgentDeliveredToCustomer сервисные колонки хранят
# уже вычтенные из amount составляющие:
#   accruals_for_sale   = комиссия Ozon (MarketplaceServiceItemFeeRevShare, < 0)
#   delivery_charge     = стоимость доставки покупателю (< 0)
#   return_delivery_charge = стоимость обратной логистики (< 0)
#   sale_commission     = комиссия за возврат после доставки (< 0)
#
# Т.е. gross_price = amount + |accruals_for_sale| + |delivery_charge|
# (сервисные суммы уже сидят внутри amount, их нужно прибавить обратно)

RETURN_OPS = {
    "OperationReturnGoodsFBSofRMS",
    "OperationReturnGoodsFBO",
    "OperationItemReturn",
    "OperationReturnGoodsFBSofMerchant",
    "OperationReturnGoodsFBSofRMSReject",
    "OperationLagTimeReturnAfterDelivery",
}

def classify(op_type: str, amount: float) -> str:
    if op_type == "OperationAgentDeliveredToCustomer":
        return "Продажи"
    if op_type in RETURN_OPS or "return" in op_type.lower() or "Return" in op_type:
        return "Возвраты"
    if "storage" in op_type.lower() or "Storage" in op_type:
        return "Хранение"
    if "logistic" in op_type.lower() or "Logistic" in op_type:
        return "Логистика"
    if "FeeRevShare" in op_type or "commission" in op_type.lower():
        return "Комиссия"
    if "penalty" in op_type.lower() or "fine" in op_type.lower() or "Penalty" in op_type:
        return "Штрафы"
    if "compensation" in op_type.lower() or "Claim" in op_type or "Correction" in op_type:
        return "Компенсации"
    return "Прочие поступления" if (amount or 0) >= 0 else "Прочие удержания"

df["category"] = df.apply(lambda r: classify(r["operation_type"], r["amount"]), axis=1)

def s(series) -> float:
    return float(series.fillna(0).sum())

# ── Блок продаж ────────────────────────────────────────────────────────────────
sales_df = df[df["operation_type"] == "OperationAgentDeliveredToCustomer"]
commission_amt  = abs(s(sales_df["accruals_for_sale"]))   # Комиссия Ozon (FeeRevShare)
logistics_amt   = abs(s(sales_df["delivery_charge"]))     # Доставка покупателю
ret_logistic    = abs(s(sales_df["return_delivery_charge"]))  # Обратная логистика
# Валовая выручка = нетто продаж + то, что уже вычтено внутри amount
sales_net       = s(sales_df["amount"])
sales_gross     = sales_net + commission_amt + logistics_amt + ret_logistic
qty_sold        = int(sales_df["quantity"].fillna(0).sum())

# ── Возвраты ──────────────────────────────────────────────────────────────────
returns_df      = df[df["category"] == "Возвраты"]
returns_amt     = abs(s(returns_df["amount"]))
qty_returned    = int(returns_df["quantity"].fillna(0).sum())

# ── Прочие удержания ──────────────────────────────────────────────────────────
storage_amt     = abs(s(df[df["category"] == "Хранение"]["amount"]))
penalty_amt     = abs(s(df[df["category"] == "Штрафы"]["amount"]))
other_neg_amt   = abs(s(df[(df["category"].isin(["Прочие удержания","Логистика"]))]["amount"]))
compensations   = s(df[df["category"] == "Компенсации"]["amount"])
other_pos_amt   = s(df[df["category"] == "Прочие поступления"]["amount"])

# ── Итог к перечислению ───────────────────────────────────────────────────────
net_total = s(df["amount"])

# ── P&L ТАБЛИЦА ───────────────────────────────────────────────────────────────
st.markdown(f"### Финансовый отчёт — {calendar.month_name[month]} {year}")
st.caption(
    "Данные из финансовых транзакций Ozon (API /v3/finance/transaction/list). "
    "Период: по дате операции с 1-го по последнее число месяца."
)

def fmt(v: float, sign=False) -> str:
    prefix = "+" if sign and v > 0 else ("−" if v < 0 else "")
    return f"{prefix}{abs(v):,.0f} ₽".replace(",", " ")

pnl_rows = [
    ("📦 Выручка брутто (цена товара)",  sales_gross,     False, False),
    ("  ├─ Комиссия Ozon",               -commission_amt, False, False),
    ("  ├─ Логистика (доставка)",         -logistics_amt,  False, False),
    ("  └─ Логистика (обратная)",         -ret_logistic,   False, False),
    ("📦 Выручка нетто (продажи)",        sales_net,       False, True),
    ("↩️ Возвраты товара",               -returns_amt,    False, False),
    ("🏭 Хранение",                       -storage_amt,    False, False),
    ("⚠️ Штрафы",                         -penalty_amt,    False, False),
    ("✅ Компенсации",                     compensations,   False, False),
    ("📎 Прочие удержания",               -other_neg_amt,  False, False),
    ("📎 Прочие поступления",             other_pos_amt,   False, False),
    ("💳 К перечислению на счёт",         net_total,       False, True),
]

pnl_df = pd.DataFrame(
    [(name, fmt(val)) for name, val, _, _ in pnl_rows],
    columns=["Статья", "Сумма, ₽"]
)
# Убираем нулевые строки (кроме итогов)
bold_idx = {i for i, (_, _, _, bold) in enumerate(pnl_rows) if bold}
zero_idx  = {i for i, (_, val, _, bold) in enumerate(pnl_rows) if val == 0 and not bold}
pnl_df = pnl_df.drop(index=list(zero_idx)).reset_index(drop=True)

st.dataframe(pnl_df, use_container_width=True, hide_index=True, height=420)

# ── KPI ───────────────────────────────────────────────────────────────────────
st.markdown("---")
k1, k2, k3, k4, k5 = st.columns(5)
k1.metric("Продано (шт.)",   qty_sold)
k2.metric("Возвратов (шт.)", qty_returned)
k3.metric("Выручка брутто",  f"{sales_gross:,.0f} ₽".replace(",", " "))
k4.metric("Комиссия + лог.", f"−{commission_amt+logistics_amt:,.0f} ₽".replace(",", " "))
k5.metric("К перечислению",  f"{net_total:,.0f} ₽".replace(",", " "))

st.markdown("---")

# ── Waterfall ─────────────────────────────────────────────────────────────────
st.markdown("#### Структура выплаты")
wf_items = [
    ("Выручка\nбрутто",     "absolute", sales_gross),
    ("Комиссия",            "relative", -commission_amt),
    ("Логистика",           "relative", -(logistics_amt + ret_logistic)),
    ("Возвраты",            "relative", -returns_amt),
]
if storage_amt:  wf_items.append(("Хранение",  "relative", -storage_amt))
if penalty_amt:  wf_items.append(("Штрафы",    "relative", -penalty_amt))
other_sum = compensations + other_pos_amt - other_neg_amt
if abs(other_sum) > 1: wf_items.append(("Прочее", "relative", other_sum))
wf_items.append(("К перечислению", "total", net_total))

labels   = [x[0] for x in wf_items]
measures = [x[1] for x in wf_items]
values   = [x[2] for x in wf_items]

fig_wf = go.Figure(go.Waterfall(
    orientation="v", measure=measures, x=labels, y=values,
    text=[f"{abs(v):,.0f}".replace(",", " ") for v in values],
    textposition="outside",
    connector={"line": {"color": "#9CA3AF"}},
    increasing={"marker": {"color": "#10B981"}},
    decreasing={"marker": {"color": "#EF4444"}},
    totals={"marker": {"color": "#F97316"}},
))
fig_wf.update_layout(margin=dict(t=40, b=10), height=400, showlegend=False)
st.plotly_chart(fig_wf, use_container_width=True)

st.markdown("---")

# ── Детализация по товарам ─────────────────────────────────────────────────────
st.markdown("#### Детализация по товарам")
if not sales_df.empty:
    grp = sales_df.groupby(["offer_id", "product_name"]).agg(
        qty         =("quantity",             "sum"),
        gross       =("amount",               lambda x: (
                          x.sum()
                          + abs(sales_df.loc[x.index, "accruals_for_sale"].fillna(0).sum())
                          + abs(sales_df.loc[x.index, "delivery_charge"].fillna(0).sum())
                          + abs(sales_df.loc[x.index, "return_delivery_charge"].fillna(0).sum())
                      )),
        commission  =("accruals_for_sale",    lambda x: abs(x.fillna(0).sum())),
        logistics   =("delivery_charge",      lambda x: abs(x.fillna(0).sum())),
        net         =("amount",               "sum"),
    ).reset_index().sort_values("net", ascending=False)

    grp = grp.rename(columns={
        "offer_id":    "Артикул",
        "product_name":"Название",
        "qty":         "Прод. шт.",
        "gross":       "Выручка брутто ₽",
        "commission":  "Комиссия ₽",
        "logistics":   "Логистика ₽",
        "net":         "Нетто ₽",
    })
    st.dataframe(
        grp, use_container_width=True, hide_index=True, height=400,
        column_config={
            "Выручка брутто ₽": st.column_config.NumberColumn(format="%.0f"),
            "Комиссия ₽":       st.column_config.NumberColumn(format="%.0f"),
            "Логистика ₽":      st.column_config.NumberColumn(format="%.0f"),
            "Нетто ₽":          st.column_config.NumberColumn(format="%.0f"),
        },
    )
else:
    st.info("Нет данных о продажах за период.")

st.markdown("---")

# ── Все типы операций ─────────────────────────────────────────────────────────
with st.expander("📋 Все типы операций (детально)"):
    by_type = (
        df.groupby(["category", "operation_type_name"])["amount"]
        .agg(["sum", "count"])
        .reset_index()
        .rename(columns={"sum": "Сумма ₽", "count": "Кол-во", "operation_type_name": "Тип операции", "category": "Категория"})
        .sort_values("Сумма ₽", ascending=False)
    )
    st.dataframe(
        by_type, use_container_width=True, hide_index=True,
        column_config={"Сумма ₽": st.column_config.NumberColumn(format="%.0f ₽")},
    )

# ── Excel-экспорт ─────────────────────────────────────────────────────────────
def to_excel() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "P&L"

    orange = PatternFill("solid", fgColor="F97316")
    light  = PatternFill("solid", fgColor="FFF7ED")
    green  = PatternFill("solid", fgColor="D1FAE5")
    bold   = Font(bold=True)
    thin   = Side(style="thin")

    ws.merge_cells("A1:B1")
    c = ws["A1"]
    c.value = f"Финансовый отчёт Ozon — {calendar.month_name[month]} {year}"
    c.font = Font(bold=True, size=13)

    ws.merge_cells("A2:B2")
    ws["A2"].value = f"Период: 01.{month:02d}.{year} — {calendar.monthrange(year,month)[1]:02d}.{month:02d}.{year}"
    ws["A2"].font = Font(italic=True, color="6B7280")

    ws.append([])  # row 3 empty

    headers = ["Статья", "Сумма, ₽"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(4, ci, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = orange
        cell.alignment = Alignment(horizontal="center")

    row_defs = [
        ("Выручка брутто (цена товара)",   sales_gross,    True),
        ("  Комиссия Ozon",                -commission_amt, False),
        ("  Логистика (доставка)",          -logistics_amt,  False),
        ("  Логистика (обратная)",          -ret_logistic,   False),
        ("Выручка нетто (продажи)",         sales_net,       True),
        ("Возвраты товара",                 -returns_amt,    False),
        ("Хранение",                        -storage_amt,    False),
        ("Штрафы",                          -penalty_amt,    False),
        ("Компенсации",                     compensations,   False),
        ("Прочие удержания",                -other_neg_amt,  False),
        ("Прочие поступления",              other_pos_amt,   False),
        ("К перечислению на счёт",          net_total,       True),
    ]

    for ri, (name, val, is_bold) in enumerate(row_defs, 5):
        if val == 0 and not is_bold:
            continue
        c1 = ws.cell(ri, 1, name)
        c2 = ws.cell(ri, 2, round(val, 2))
        c2.number_format = '#,##0.00 ₽'
        if is_bold:
            c1.font = Font(bold=True)
            c2.font = Font(bold=True)
            c1.fill = light
            c2.fill = light
        if name.startswith("К перечислению"):
            c1.fill = green
            c2.fill = green

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20

    # Sheet 2: by product
    ws2 = wb.create_sheet("По товарам")
    if not sales_df.empty:
        prod_cols = ["Артикул", "Название", "Прод. шт.", "Выручка брутто ₽", "Комиссия ₽", "Логистика ₽", "Нетто ₽"]
        for ci, h in enumerate(prod_cols, 1):
            c = ws2.cell(1, ci, h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = orange
        for ri, row in enumerate(grp.itertuples(index=False), 2):
            for ci, val in enumerate(row, 1):
                ws2.cell(ri, ci, val)
        for col in ws2.columns:
            mx = max((len(str(c.value or "")) for c in col), default=8)
            ws2.column_dimensions[get_column_letter(col[0].column)].width = min(mx + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

st.download_button(
    "⬇️ Скачать P&L Excel",
    data=to_excel(),
    file_name=f"ozon_pnl_{year}_{month:02d}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)
